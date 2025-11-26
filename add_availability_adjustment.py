#!/usr/bin/env python3
"""
Add Availability_Adj to Ranking_Score
- 90%+ = +2 points (very reliable)
- 80-89% = +1 point (reliable)
- 70-79% = 0 points (neutral)
- 60-69% = -2 points (concern)
- <60% = -4 points (major red flag)
"""

import pandas as pd
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

print("\nCalculating Availability_Adj...")

# Calculate Availability_Adj
availability_adjs = []
for idx, row in df.iterrows():
    avail_pct = row['Career_Availability_Pct']
    
    if pd.notna(avail_pct):
        if avail_pct >= 90:
            adj = 2
        elif avail_pct >= 80:
            adj = 1
        elif avail_pct >= 70:
            adj = 0
        elif avail_pct >= 60:
            adj = -2
        else:  # <60%
            adj = -4
        availability_adjs.append(adj)
    else:
        # No data - neutral
        availability_adjs.append(0)

df['Availability_Adj'] = availability_adjs

# Recalculate Ranking_Score (add Availability_Adj)
print("Recalculating Ranking_Score...")
ranking_scores = []
for idx, row in df.iterrows():
    base = row['Weighted_convBA+'] if pd.notna(row['Weighted_convBA+']) else 0
    defense = row['Def_Spectrum_Pts']
    age = row['Age_Factor']
    manual = row['Manual_Adj']
    avail_adj = row['Availability_Adj']
    
    if base > 0:  # Only rank players with stats
        score = round(base + defense + age + manual + avail_adj, 1)
        ranking_scores.append(score)
    else:
        ranking_scores.append(None)

df['Ranking_Score'] = ranking_scores

# Sort by Ranking_Score (descending)
df = df.sort_values('Ranking_Score', ascending=False, na_position='last')

# Reorder columns - add Availability_Adj after Manual_Adj
new_order = [
    'PID',
    'FirstName',
    'LastName',
    'Position',
    'Age',
    'PA',
    'HR',
    'Career_Games_Played',
    'Career_Games_Possible',
    'Career_Availability_Pct',
    'Win_Pct',
    'Career_Wins',
    'Career_Losses',
    'convBA_F25',
    'convBA_S25',
    'convBA_W25',
    'convBA+_F25',
    'convBA+_S25',
    'convBA+_W25',
    'Weighted_convBA+',
    'Def_Spectrum_Pts',
    'Age_Factor',
    'Manual_Adj',
    'Availability_Adj',
    'Ranking_Score',
    'Availability'
]

df = df[new_order]

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns (N, O, P)
for col_idx in [14, 15, 16]:  # convBA_F25, S25, W25
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column (K)
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)  # Win_Pct column
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Added: Availability_Adj")
print(f"    90%+ = +2 points")
print(f"    80-89% = +1 point")
print(f"    70-79% = 0 points")
print(f"    60-69% = -2 points")
print(f"    <60% = -4 points")
print(f"  Updated: Ranking_Score")
print(f"  Re-sorted by Ranking_Score")
print(f"  Backup: {backup_file}")

print(f"\nTop 10 with Availability_Adj:")
display_cols = ['FirstName', 'LastName', 'Career_Availability_Pct', 'Availability_Adj', 'Ranking_Score']
print(df[display_cols].head(10).to_string(index=False))

print(f"\nBiggest movers from availability adjustment:")
print("\nPlayers who got +2 (90%+ availability) in top 20:")
top20 = df.head(20)
boosted = top20[top20['Availability_Adj'] == 2]
if len(boosted) > 0:
    print(boosted[['FirstName', 'LastName', 'Career_Availability_Pct', 'Availability_Adj', 'Ranking_Score']].to_string(index=False))

print(f"\nPlayers who got penalized (negative adj) in top 30:")
top30 = df.head(30)
penalized = top30[top30['Availability_Adj'] < 0]
if len(penalized) > 0:
    print(penalized[['FirstName', 'LastName', 'Career_Availability_Pct', 'Availability_Adj', 'Ranking_Score']].to_string(index=False))

