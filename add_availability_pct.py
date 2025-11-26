#!/usr/bin/env python3
"""
Add Availability % - percentage of team games played in F25
Example: Team played 20 games, player played 18 = 90%
"""

import pandas as pd
import sqlite3
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'
DATABASE_PATH = 'softball_stats.db'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# Connect to database
print("\nCalculating F25 availability percentage...")
conn = sqlite3.connect(DATABASE_PATH)

def get_f25_availability_pct(pid):
    """Calculate percentage of team games played in F25"""
    # Find player's F25 team
    team_query = """
        SELECT DISTINCT b.TeamNumber, t.LongTeamName
        FROM batting_stats b
        INNER JOIN Teams t ON b.TeamNumber = t.TeamNumber
        WHERE b.PlayerNumber = ?
        AND t.LongTeamName LIKE '% F25'
        LIMIT 1
    """
    
    cursor = conn.cursor()
    cursor.execute(team_query, (pid,))
    team_result = cursor.fetchone()
    
    if not team_result:
        return None  # Didn't play F25
    
    team_number = team_result[0]
    
    # Count total games the team played in F25
    total_games_query = """
        SELECT COUNT(DISTINCT GameNumber)
        FROM game_stats
        WHERE TeamNumber = ?
    """
    cursor.execute(total_games_query, (team_number,))
    total_games = cursor.fetchone()[0]
    
    # Count games player appeared in
    player_games_query = """
        SELECT COUNT(DISTINCT GameNumber)
        FROM batting_stats
        WHERE PlayerNumber = ?
        AND TeamNumber = ?
    """
    cursor.execute(player_games_query, (pid, team_number))
    player_games = cursor.fetchone()[0]
    
    if total_games > 0:
        availability_pct = round((player_games / total_games) * 100, 1)
        return availability_pct
    else:
        return None

# Calculate for each player
availability_pcts = []
games_played_f25 = []
team_games_f25 = []

for idx, row in df.iterrows():
    pid = row['PID']
    
    if pd.notna(pid):
        pid = int(pid)
        
        # Get player's F25 team and game counts
        cursor = conn.cursor()
        
        # Find player's F25 team
        cursor.execute("""
            SELECT DISTINCT b.TeamNumber
            FROM batting_stats b
            INNER JOIN Teams t ON b.TeamNumber = t.TeamNumber
            WHERE b.PlayerNumber = ?
            AND t.LongTeamName LIKE '% F25'
            LIMIT 1
        """, (pid,))
        team_result = cursor.fetchone()
        
        if team_result:
            team_number = team_result[0]
            
            # Total team games
            cursor.execute("""
                SELECT COUNT(DISTINCT GameNumber)
                FROM game_stats
                WHERE TeamNumber = ?
            """, (team_number,))
            total_games = cursor.fetchone()[0]
            
            # Player games
            cursor.execute("""
                SELECT COUNT(DISTINCT GameNumber)
                FROM batting_stats
                WHERE PlayerNumber = ?
                AND TeamNumber = ?
            """, (pid, team_number))
            player_games = cursor.fetchone()[0]
            
            if total_games > 0:
                avail_pct = round((player_games / total_games) * 100, 1)
                availability_pcts.append(avail_pct)
                games_played_f25.append(player_games)
                team_games_f25.append(total_games)
            else:
                availability_pcts.append(None)
                games_played_f25.append(0)
                team_games_f25.append(0)
        else:
            # Didn't play F25
            availability_pcts.append(None)
            games_played_f25.append(0)
            team_games_f25.append(0)
    else:
        availability_pcts.append(None)
        games_played_f25.append(0)
        team_games_f25.append(0)

conn.close()

# Add new columns
df['Games_Played_F25'] = games_played_f25
df['Team_Games_F25'] = team_games_f25
df['Availability_Pct'] = availability_pcts

# Reorder columns - add after PA/HR
new_order = [
    'PID',
    'FirstName',
    'LastName',
    'Position',
    'Age',
    'PA',
    'HR',
    'Games_Played_F25',
    'Team_Games_F25',
    'Availability_Pct',
    'Win_Pct',
    'Career_Wins',
    'Career_Losses',
    'convBA_F25',
    'convBA_S25',
    'convBA_W25',
    'convBA+_F25',
    'convBA+_S25',
    'convBA+_W25',
    'Weighted_convBA+',
    'Def_Spectrum_Pts',
    'Age_Factor',
    'Manual_Adj',
    'Ranking_Score',
    'Availability'
]

df = df[new_order]

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns (N, O, P - adjusted for new columns)
for col_idx in [14, 15, 16]:  # convBA_F25, S25, W25
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column (K)
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)  # Win_Pct column
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Added: Games_Played_F25, Team_Games_F25, Availability_Pct")
print(f"  Shows percentage of team games played in F25")
print(f"  Backup: {backup_file}")

print(f"\nTop 10 with Availability %:")
display_cols = ['FirstName', 'LastName', 'Games_Played_F25', 'Team_Games_F25', 'Availability_Pct', 'Ranking_Score']
print(df[display_cols].head(10).to_string(index=False))

print(f"\nPerfect Attendance (100%):")
perfect = df[df['Availability_Pct'] == 100.0]
if len(perfect) > 0:
    print(f"  {len(perfect)} players with 100% availability")
    print(perfect[['FirstName', 'LastName', 'Games_Played_F25', 'Team_Games_F25']].head(10).to_string(index=False))

