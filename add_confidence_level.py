#!/usr/bin/env python3
"""
Add Confidence Level to rankings based on:
1. Data Recency (40%): How recent is their most recent season?
2. Sample Size (30%): Total PA across seasons used
3. Number of Seasons (20%): More seasons = more data
4. Consistency (10%): Low variance = more predictable

Final: A (90-100%), B (80-89%), C (70-79%), D (60-69%), F (<60%)
"""

import pandas as pd
import sqlite3
from datetime import datetime
import shutil
from openpyxl import load_workbook
import numpy as np

EXCEL_FILE = 'w26rankings.xlsx'
DATABASE_PATH = 'softball_stats.db'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# Connect to database
print("\nCalculating confidence levels...")
conn = sqlite3.connect(DATABASE_PATH)

def get_season_recency_score(season_str):
    """Score based on how recent the season is"""
    if not season_str:
        return 0
    
    # F25, W25 (current): 100%
    if season_str in ['F25', 'W25']:
        return 1.0
    # S25 (recent): 80%
    elif season_str == 'S25':
        return 0.8
    # F24, W24 (one year ago): 60%
    elif season_str in ['F24', 'W24']:
        return 0.6
    # S24 or older: 40%
    else:
        return 0.4

def get_pa_for_seasons(pid, seasons_used_str):
    """Get total PA for the seasons used"""
    if not seasons_used_str or not pid:
        return 0
    
    seasons = [s.strip() for s in seasons_used_str.split(',')]
    total_pa = 0
    
    cursor = conn.cursor()
    for season in seasons:
        query = """
            SELECT SUM(PA) as PA
            FROM batting_stats bs
            INNER JOIN Teams t ON bs.TeamNumber = t.TeamNumber
            WHERE bs.PlayerNumber = ?
            AND t.LongTeamName LIKE ?
        """
        cursor.execute(query, (pid, f'% {season}'))
        result = cursor.fetchone()
        if result and result[0]:
            total_pa += result[0]
    
    return total_pa

def get_convba_plus_variance(pid, seasons_used_str):
    """Calculate variance in convBA+ across seasons"""
    if not seasons_used_str or not pid:
        return 0
    
    seasons = [s.strip() for s in seasons_used_str.split(',')]
    if len(seasons) < 2:
        return 0  # Can't calculate variance with 1 season
    
    convba_plus_values = []
    cursor = conn.cursor()
    
    for season in seasons:
        # Get convBA+ for this season
        query = """
            SELECT 
                SUM(PA) as PA,
                SUM(H) as H,
                SUM(BB) as BB,
                SUM(SF) as SF,
                SUM([2B]) as Doubles,
                SUM([3B]) as Triples,
                SUM(HR) as HR
            FROM batting_stats bs
            INNER JOIN Teams t ON bs.TeamNumber = t.TeamNumber
            WHERE bs.PlayerNumber = ?
            AND t.LongTeamName LIKE ?
        """
        cursor.execute(query, (pid, f'% {season}'))
        result = cursor.fetchone()
        
        if result and result[0]:
            pa, h, bb, sf, doubles, triples, hr = result
            if pa > 0:
                total_bases = h + doubles + (2 * triples) + (3 * hr)
                conv_ba = (((4*(h+bb)+total_bases)/pa)/0.305*0.25)/10
                
                # Get league average for this season
                league_query = """
                    SELECT 
                        SUM(PA) as PA,
                        SUM(H) as H,
                        SUM(BB) as BB,
                        SUM([2B]) as Doubles,
                        SUM([3B]) as Triples,
                        SUM(HR) as HR
                    FROM batting_stats bs
                    INNER JOIN Teams t ON bs.TeamNumber = t.TeamNumber
                    WHERE t.LongTeamName LIKE ?
                """
                cursor.execute(league_query, (f'% {season}',))
                league_row = cursor.fetchone()
                
                if league_row and league_row[0]:
                    l_pa, l_h, l_bb, l_doubles, l_triples, l_hr = league_row
                    l_tb = l_h + l_doubles + (2 * l_triples) + (3 * l_hr)
                    league_conv_ba = (((4*(l_h+l_bb)+l_tb)/l_pa)/0.305*0.25)/10
                    
                    conv_ba_plus = (conv_ba / league_conv_ba) * 100
                    convba_plus_values.append(conv_ba_plus)
    
    if len(convba_plus_values) >= 2:
        return np.std(convba_plus_values)
    return 0

# Calculate confidence for each player
confidence_scores = []
confidence_grades = []
recency_scores = []
sample_size_scores = []
num_seasons_scores = []
consistency_scores = []

for idx, row in df.iterrows():
    pid = row['PID']
    seasons_used = row.get('Seasons_Used', '')
    
    if pd.notna(pid) and seasons_used:
        pid = int(pid)
        
        # 1. Recency Score (40% weight) - based on most recent season
        most_recent = seasons_used.split(',')[0].strip() if seasons_used else ''
        recency = get_season_recency_score(most_recent)
        recency_scores.append(recency)
        
        # 2. Sample Size Score (30% weight) - total PA
        total_pa = get_pa_for_seasons(pid, seasons_used)
        if total_pa >= 150:
            sample_size = 1.0
        elif total_pa >= 100:
            sample_size = 0.75
        elif total_pa >= 50:
            sample_size = 0.5
        else:
            sample_size = 0.25
        sample_size_scores.append(sample_size)
        
        # 3. Number of Seasons Score (20% weight)
        num_seasons = len(seasons_used.split(','))
        if num_seasons >= 3:
            seasons_score = 1.0
        elif num_seasons == 2:
            seasons_score = 0.67
        else:
            seasons_score = 0.33
        num_seasons_scores.append(seasons_score)
        
        # 4. Consistency Score (10% weight) - variance in convBA+
        variance = get_convba_plus_variance(pid, seasons_used)
        if num_seasons == 1:
            consistency = 0.75  # Can't measure variance, give benefit of doubt
        elif variance < 15:
            consistency = 1.0   # Very consistent
        elif variance < 30:
            consistency = 0.75  # Moderate variance
        else:
            consistency = 0.5   # High variance
        consistency_scores.append(consistency)
        
        # Calculate overall confidence (weighted average)
        confidence = (recency * 0.40 + 
                     sample_size * 0.30 + 
                     seasons_score * 0.20 + 
                     consistency * 0.10)
        
        confidence_pct = round(confidence * 100, 1)
        confidence_scores.append(confidence_pct)
        
        # Assign grade
        if confidence_pct >= 90:
            grade = 'A'
        elif confidence_pct >= 80:
            grade = 'B'
        elif confidence_pct >= 70:
            grade = 'C'
        elif confidence_pct >= 60:
            grade = 'D'
        else:
            grade = 'F'
        confidence_grades.append(grade)
    else:
        # No data
        confidence_scores.append(0)
        confidence_grades.append('F')
        recency_scores.append(0)
        sample_size_scores.append(0)
        num_seasons_scores.append(0)
        consistency_scores.append(0)

conn.close()

# Add confidence columns
df['Confidence'] = confidence_grades
df['Confidence_Pct'] = confidence_scores

# Reorder columns - add Confidence after Seasons_Used
cols = df.columns.tolist()
seasons_idx = cols.index('Seasons_Used')
cols.insert(seasons_idx + 1, cols.pop(cols.index('Confidence')))
cols.insert(seasons_idx + 2, cols.pop(cols.index('Confidence_Pct')))
df = df[cols]

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns
for col_idx in [14, 15, 16]:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Added: Confidence (grade A-F) and Confidence_Pct (0-100)")
print(f"  Formula: Recency 40% + Sample Size 30% + Seasons 20% + Consistency 10%")
print(f"  Backup: {backup_file}")

print(f"\nTop 20 with confidence levels:")
display_cols = ['FirstName', 'LastName', 'Seasons_Used', 'Confidence', 'Confidence_Pct', 'Ranking_Score']
print(df[display_cols].head(20).to_string(index=False))

print(f"\nConfidence distribution in top 50:")
top50 = df.head(50)
conf_dist = top50['Confidence'].value_counts().sort_index()
print(conf_dist.to_string())

print(f"\nHigh-confidence elites (A grade, top 30):")
top30_a = df.head(30)[df.head(30)['Confidence'] == 'A']
if len(top30_a) > 0:
    print(top30_a[['FirstName', 'LastName', 'Seasons_Used', 'Confidence_Pct', 'Ranking_Score']].to_string(index=False))

print(f"\nLow-confidence high ranks (D/F grade, top 30):")
top30_low = df.head(30)[df.head(30)['Confidence'].isin(['D', 'F'])]
if len(top30_low) > 0:
    print(top30_low[['FirstName', 'LastName', 'Seasons_Used', 'Confidence', 'Confidence_Pct', 'Ranking_Score']].to_string(index=False))

