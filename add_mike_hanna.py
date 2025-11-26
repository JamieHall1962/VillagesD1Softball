#!/usr/bin/env python3
"""
Add Mike Hanna back to rankings (PID 530, Position RF)
He was incorrectly converted to Mike Hage
"""

import pandas as pd
import sqlite3
from datetime import datetime
import shutil
from openpyxl import load_workbook
import numpy as np

EXCEL_FILE = 'w26rankings.xlsx'
DATABASE_PATH = 'softball_stats.db'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# Connect to database
print("\nGetting Mike Hanna data from database...")
conn = sqlite3.connect(DATABASE_PATH)

pid = 530
position = 'RF'

# Get basic info
cursor = conn.cursor()
cursor.execute("SELECT FirstName, LastName FROM People WHERE PersonNumber = ?", (pid,))
result = cursor.fetchone()
first_name = result[0] if result else "Mike"
last_name = result[1] if result else "Hanna"

print(f"  Found: {first_name} {last_name}, PID {pid}")

# Get age from w26reg.xlsx
try:
    reg_df = pd.read_excel('w26reg.xlsx', sheet_name='full_time_players')
    age_row = reg_df[reg_df['PersonNumber'] == pid]
    age = int(age_row['Age'].iloc[0]) if len(age_row) > 0 else None
except:
    age = None

# Calculate age factor
if age:
    if age < 60:
        age_factor = 5
    elif age <= 64:
        age_factor = 3
    elif age <= 69:
        age_factor = 1
    elif age <= 74:
        age_factor = 0
    elif age <= 79:
        age_factor = -2
    else:
        age_factor = -4
else:
    age_factor = 0

# Get all seasons with stats
query = """
    SELECT t.LongTeamName,
           SUM(PA) as PA,
           SUM(H) as H,
           SUM(BB) as BB,
           SUM(SF) as SF,
           SUM(OE) as OE,
           SUM([2B]) as Doubles,
           SUM([3B]) as Triples,
           SUM(HR) as HR
    FROM batting_stats bs
    INNER JOIN Teams t ON bs.TeamNumber = t.TeamNumber
    WHERE bs.PlayerNumber = ?
    GROUP BY bs.TeamNumber
    ORDER BY t.LongTeamName DESC
"""

cursor.execute(query, (pid,))
seasons_data = cursor.fetchall()

# Calculate convBA+ for each season and get last 3
def get_season_sort_value(season_str):
    if not season_str or len(season_str) < 3:
        return 0
    season_type = season_str[0]
    try:
        year = int(season_str[1:])
    except:
        return 0
    season_order = {'F': 3, 'S': 2, 'W': 1}.get(season_type, 0)
    return year * 10 + season_order

season_convba_plus = []
for season_row in seasons_data:
    team_name = season_row[0]
    pa = season_row[1]
    h = season_row[2]
    bb = season_row[3]
    sf = season_row[4]
    oe = season_row[5]
    doubles = season_row[6]
    triples = season_row[7]
    hr = season_row[8]
    
    if pa > 0:
        total_bases = h + doubles + (2 * triples) + (3 * hr)
        conv_ba = (((4*(h+bb)+total_bases)/pa)/0.305*0.25)/10
        
        season_code = team_name.split()[-1] if team_name else None
        if season_code:
            # Get league average
            league_query = """
                SELECT SUM(PA) as PA, SUM(H) as H, SUM(BB) as BB,
                       SUM([2B]) as Doubles, SUM([3B]) as Triples, SUM(HR) as HR
                FROM batting_stats bs
                INNER JOIN Teams t ON bs.TeamNumber = t.TeamNumber
                WHERE t.LongTeamName LIKE ?
            """
            cursor.execute(league_query, (f'% {season_code}',))
            league_row = cursor.fetchone()
            
            if league_row and league_row[0]:
                l_pa, l_h, l_bb, l_doubles, l_triples, l_hr = league_row
                l_tb = l_h + l_doubles + (2 * l_triples) + (3 * l_hr)
                league_conv_ba = (((4*(l_h+l_bb)+l_tb)/l_pa)/0.305*0.25)/10
                
                conv_ba_plus = (conv_ba / league_conv_ba) * 100
                season_convba_plus.append({
                    'season': season_code,
                    'convBA+': conv_ba_plus,
                    'convBA': conv_ba,
                    'sort_value': get_season_sort_value(season_code),
                    'PA': pa,
                    'HR': hr
                })

# Sort and take last 3
season_convba_plus.sort(key=lambda x: x['sort_value'], reverse=True)
last_3_seasons = season_convba_plus[:3]

# Get F25, S25, W25 values for display
convba_f25 = next((s['convBA'] for s in season_convba_plus if s['season'] == 'F25'), 0)
convba_s25 = next((s['convBA'] for s in season_convba_plus if s['season'] == 'S25'), None)
convba_w25 = next((s['convBA'] for s in season_convba_plus if s['season'] == 'W25'), None)

convba_plus_f25 = next((s['convBA+'] for s in season_convba_plus if s['season'] == 'F25'), 0)
convba_plus_s25 = next((s['convBA+'] for s in season_convba_plus if s['season'] == 'S25'), None)
convba_plus_w25 = next((s['convBA+'] for s in season_convba_plus if s['season'] == 'W25'), None)

# Calculate weighted average from last 3 seasons
if last_3_seasons:
    base_weights = [0.50, 0.30, 0.20]
    total_weight = sum(base_weights[:len(last_3_seasons)])
    weighted_sum = sum(s['convBA+'] * (base_weights[i] / total_weight) 
                      for i, s in enumerate(last_3_seasons))
    weighted_convba_plus = round(weighted_sum, 1)
    seasons_used = ', '.join([s['season'] for s in last_3_seasons])
    pa = last_3_seasons[0]['PA'] if last_3_seasons else 0
    hr = last_3_seasons[0]['HR'] if last_3_seasons else 0
else:
    weighted_convba_plus = 0
    seasons_used = ''
    pa = 0
    hr = 0

# Calculate career availability
cursor.execute("""
    SELECT DISTINCT b.TeamNumber
    FROM batting_stats b
    WHERE b.PlayerNumber = ?
    ORDER BY b.TeamNumber DESC
    LIMIT 20
""", (pid,))
teams = cursor.fetchall()

total_games_played = 0
total_games_possible = 0

for team in teams:
    team_number = team[0]
    cursor.execute("SELECT COUNT(DISTINCT GameNumber) FROM game_stats WHERE TeamNumber = ?", (team_number,))
    team_games = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(DISTINCT GameNumber) FROM batting_stats WHERE PlayerNumber = ? AND TeamNumber = ?", 
                  (pid, team_number))
    player_games = cursor.fetchone()[0]
    total_games_played += player_games
    total_games_possible += team_games

career_availability_pct = round((total_games_played / total_games_possible) * 100, 1) if total_games_possible > 0 else None

# Calculate availability adjustment
if career_availability_pct:
    if career_availability_pct >= 90:
        avail_adj = 2
    elif career_availability_pct >= 80:
        avail_adj = 1
    elif career_availability_pct >= 70:
        avail_adj = 0
    elif career_availability_pct >= 60:
        avail_adj = -2
    else:
        avail_adj = -4
else:
    avail_adj = 0

# Calculate win %
cursor.execute("""
    SELECT g.Runs, g.OppRuns
    FROM batting_stats b
    INNER JOIN game_stats g ON b.TeamNumber = g.TeamNumber AND b.GameNumber = g.GameNumber
    WHERE b.PlayerNumber = ?
    GROUP BY g.TeamNumber, g.GameNumber
""", (pid,))
games = cursor.fetchall()

wins = sum(1 for g in games if g[0] > g[1])
losses = sum(1 for g in games if g[0] < g[1])
ties = sum(1 for g in games if g[0] == g[1])

win_pct = round(wins / (wins + losses), 3) if (wins + losses) > 0 else 0.000

# Calculate confidence
most_recent = last_3_seasons[0]['season'] if last_3_seasons else ''
if most_recent in ['F25', 'W25']:
    recency = 1.0
elif most_recent == 'S25':
    recency = 0.8
elif most_recent in ['F24', 'W24']:
    recency = 0.6
else:
    recency = 0.4

total_pa = sum(s['PA'] for s in last_3_seasons)
if total_pa >= 150:
    sample_size = 1.0
elif total_pa >= 100:
    sample_size = 0.75
elif total_pa >= 50:
    sample_size = 0.5
else:
    sample_size = 0.25

num_seasons = len(last_3_seasons)
if num_seasons >= 3:
    seasons_score = 1.0
elif num_seasons == 2:
    seasons_score = 0.67
else:
    seasons_score = 0.33

if num_seasons == 1:
    consistency = 0.75
elif num_seasons >= 2:
    variance = np.std([s['convBA+'] for s in last_3_seasons])
    if variance < 15:
        consistency = 1.0
    elif variance < 30:
        consistency = 0.75
    else:
        consistency = 0.5
else:
    consistency = 0.75

confidence_pct = round((recency * 0.40 + sample_size * 0.30 + seasons_score * 0.20 + consistency * 0.10) * 100, 1)
if confidence_pct >= 90:
    confidence = 'A'
elif confidence_pct >= 80:
    confidence = 'B'
elif confidence_pct >= 70:
    confidence = 'C'
elif confidence_pct >= 60:
    confidence = 'D'
else:
    confidence = 'F'

conn.close()

# Defensive spectrum points
def_spectrum_pts = 1.5  # RF

# Manual adjustment
manual_adj = 0

# Calculate ranking score
ranking_score = round(weighted_convba_plus + def_spectrum_pts + age_factor + manual_adj + avail_adj, 1) if weighted_convba_plus > 0 else None

# Create new row
new_row = {
    'PID': pid,
    'FirstName': first_name,
    'LastName': last_name,
    'Position': position,
    'Age': age,
    'PA': pa,
    'HR': hr,
    'Career_Games_Played': total_games_played,
    'Career_Games_Possible': total_games_possible,
    'Career_Availability_Pct': career_availability_pct,
    'Win_Pct': win_pct,
    'Career_Wins': wins,
    'Career_Losses': losses,
    'convBA_F25': convba_f25,
    'convBA_S25': convba_s25,
    'convBA_W25': convba_w25,
    'convBA+_F25': convba_plus_f25,
    'convBA+_S25': convba_plus_s25,
    'convBA+_W25': convba_plus_w25,
    'Weighted_convBA+': weighted_convba_plus,
    'Seasons_Used': seasons_used,
    'Confidence': confidence,
    'Confidence_Pct': confidence_pct,
    'Def_Spectrum_Pts': def_spectrum_pts,
    'Age_Factor': age_factor,
    'Manual_Adj': manual_adj,
    'Availability_Adj': avail_adj,
    'Ranking_Score': ranking_score,
    'Availability': ''  # Will get from registration if needed
}

print(f"\nMike Hanna stats:")
print(f"  Position: {position}")
print(f"  Age: {age}, Age Factor: {age_factor}")
print(f"  PA: {pa}, HR: {hr}")
print(f"  Seasons Used: {seasons_used}")
print(f"  Weighted convBA+: {weighted_convba_plus}")
print(f"  Career Availability: {career_availability_pct}% (Adj: {avail_adj})")
print(f"  Win %: {win_pct} ({wins}-{losses})")
print(f"  Confidence: {confidence} ({confidence_pct}%)")
print(f"  Ranking Score: {ranking_score}")

# Add to dataframe
df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

# Sort by Ranking_Score
df = df.sort_values('Ranking_Score', ascending=False, na_position='last')

print(f"\nAdded Mike Hanna to rankings")
print(f"  Total players: {len(df)}")

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns
for col_idx in [14, 15, 16]:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Mike Hanna added successfully")
print(f"  Backup: {backup_file}")

# Show his rank
hanna_rank = df[df['LastName'] == 'Hanna'].index[0] + 1
print(f"\nMike Hanna ranked: #{hanna_rank}")

