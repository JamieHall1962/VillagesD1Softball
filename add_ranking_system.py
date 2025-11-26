#!/usr/bin/env python3
"""
Add ranking system:
1. Weighted convBA+ (F25: 50%, S25: 30%, W25: 20%)
2. Defensive spectrum points (SS=10, LC=9, RC=8... C=0)
3. Age factor (younger = bonus)
4. Manual adjustment column (for defense/speed/intangibles)
5. Final Ranking Score
"""

import pandas as pd
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# Defensive spectrum points (SS most valuable, C least valuable)
DEFENSIVE_SPECTRUM = {
    'SS': 10,
    'LC': 9,
    'RC': 8,
    'MI': 7,
    'LF': 6,
    '3B': 5,
    '2B': 4,
    'P': 3,
    'RF': 2,
    '1B': 1,
    'C': 0
}

print("\nCalculating ranking components...")

# 1. Calculate weighted convBA+ (F25: 50%, S25: 30%, W25: 20%)
weighted_convba_plus = []
for idx, row in df.iterrows():
    f25 = row['convBA+_F25'] if pd.notna(row['convBA+_F25']) else None
    s25 = row['convBA+_S25'] if pd.notna(row['convBA+_S25']) else None
    w25 = row['convBA+_W25'] if pd.notna(row['convBA+_W25']) else None
    
    # Calculate weighted average (only use seasons they played)
    total_weight = 0
    weighted_sum = 0
    
    if f25 is not None:
        weighted_sum += f25 * 0.50
        total_weight += 0.50
    if s25 is not None:
        weighted_sum += s25 * 0.30
        total_weight += 0.30
    if w25 is not None:
        weighted_sum += w25 * 0.20
        total_weight += 0.20
    
    if total_weight > 0:
        weighted_avg = round(weighted_sum / total_weight, 1)
        weighted_convba_plus.append(weighted_avg)
    else:
        weighted_convba_plus.append(None)

df['Weighted_convBA+'] = weighted_convba_plus

# 2. Defensive spectrum points
def_points = []
for idx, row in df.iterrows():
    pos = row['Position']
    if pd.notna(pos) and pos in DEFENSIVE_SPECTRUM:
        def_points.append(DEFENSIVE_SPECTRUM[pos])
    else:
        def_points.append(0)  # Unknown position

df['Def_Spectrum_Pts'] = def_points

# 3. Age factor - need to get age from registration file
print("Loading age data from w26reg.xlsx...")
try:
    reg_df = pd.read_excel('w26reg.xlsx', sheet_name='full_time_players')
    age_lookup = dict(zip(reg_df['PersonNumber'], reg_df['Age']))
    
    ages = []
    age_factors = []
    for idx, row in df.iterrows():
        pid = row['PID']
        if pd.notna(pid) and pid in age_lookup:
            age = age_lookup[pid]
            ages.append(age)
            # Age factor: 60-65 = +3, 66-70 = +2, 71-75 = +1, 76-80 = 0, 81+ = -1
            if age <= 65:
                age_factors.append(3)
            elif age <= 70:
                age_factors.append(2)
            elif age <= 75:
                age_factors.append(1)
            elif age <= 80:
                age_factors.append(0)
            else:
                age_factors.append(-1)
        else:
            ages.append(None)
            age_factors.append(0)
    
    df['Age'] = ages
    df['Age_Factor'] = age_factors
except Exception as e:
    print(f"  Warning: Could not load age data: {e}")
    df['Age'] = None
    df['Age_Factor'] = 0

# 4. Manual adjustment column (empty for you to fill in)
df['Manual_Adj'] = 0

# 5. Calculate Final Ranking Score
# Formula: Weighted_convBA+ + Def_Spectrum_Pts + Age_Factor + Manual_Adj
ranking_scores = []
for idx, row in df.iterrows():
    base = row['Weighted_convBA+'] if pd.notna(row['Weighted_convBA+']) else 0
    defense = row['Def_Spectrum_Pts']
    age = row['Age_Factor']
    manual = row['Manual_Adj']
    
    if base > 0:  # Only rank players with stats
        score = round(base + defense + age + manual, 1)
        ranking_scores.append(score)
    else:
        ranking_scores.append(None)

df['Ranking_Score'] = ranking_scores

# Sort by Ranking_Score (descending)
df = df.sort_values('Ranking_Score', ascending=False, na_position='last')

# Reorder columns for better flow
new_order = [
    'PID',
    'FirstName',
    'LastName',
    'Position',
    'Age',
    'PA',
    'HR',
    'convBA_F25',
    'convBA_S25',
    'convBA_W25',
    'convBA+_F25',
    'convBA+_S25',
    'convBA+_W25',
    'Weighted_convBA+',
    'Def_Spectrum_Pts',
    'Age_Factor',
    'Manual_Adj',
    'Ranking_Score',
    'Availability'
]

df = df[new_order]

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply convBA formatting
print("Re-applying .000 format to convBA columns...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns (H, I, J)
for col_idx in [8, 9, 10]:  # convBA_F25, S25, W25
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Added columns:")
print(f"    - Weighted_convBA+ (F25: 50%, S25: 30%, W25: 20%)")
print(f"    - Def_Spectrum_Pts (SS=10 down to C=0)")
print(f"    - Age_Factor (+3 to -1 based on age)")
print(f"    - Manual_Adj (empty - for you to adjust)")
print(f"    - Ranking_Score (sum of all factors)")
print(f"  Sorted by Ranking_Score (highest to lowest)")
print(f"  Backup: {backup_file}")
print(f"\nTop 10 players:")
print(df[['FirstName', 'LastName', 'Position', 'Weighted_convBA+', 'Def_Spectrum_Pts', 'Age_Factor', 'Ranking_Score']].head(10).to_string(index=False))

