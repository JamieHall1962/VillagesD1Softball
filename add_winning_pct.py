#!/usr/bin/env python3
"""
Add lifetime winning percentage for each player
"""

import pandas as pd
import sqlite3
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'
DATABASE_PATH = 'softball_stats.db'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# Connect to database
print("\nCalculating lifetime winning percentage...")
conn = sqlite3.connect(DATABASE_PATH)

def get_lifetime_winning_pct(pid):
    """Calculate lifetime winning percentage for a player"""
    query = """
        SELECT 
            g.TeamNumber,
            g.GameNumber,
            g.Runs,
            g.OppRuns
        FROM batting_stats b
        INNER JOIN game_stats g ON b.TeamNumber = g.TeamNumber AND b.GameNumber = g.GameNumber
        WHERE b.PlayerNumber = ?
        GROUP BY g.TeamNumber, g.GameNumber
    """
    
    cursor = conn.cursor()
    cursor.execute(query, (pid,))
    games = cursor.fetchall()
    
    if not games:
        return None, 0, 0, 0
    
    wins = 0
    losses = 0
    ties = 0
    
    for game in games:
        runs = game[2]      # Runs
        opp_runs = game[3]  # OppRuns
        
        if runs > opp_runs:
            wins += 1
        elif runs < opp_runs:
            losses += 1
        else:
            ties += 1
    
    total_games = wins + losses + ties
    
    if (wins + losses) > 0:
        win_pct = round(wins / (wins + losses), 3)  # Ties not counted in win pct
        return win_pct, wins, losses, ties
    else:
        return 0.000, wins, losses, ties

# Calculate for each player
win_pcts = []
career_wins = []
career_losses = []
career_ties = []

for idx, row in df.iterrows():
    pid = row['PID']
    
    if pd.notna(pid):
        pid = int(pid)
        win_pct, wins, losses, ties = get_lifetime_winning_pct(pid)
        win_pcts.append(win_pct)
        career_wins.append(wins)
        career_losses.append(losses)
        career_ties.append(ties)
    else:
        win_pcts.append(None)
        career_wins.append(0)
        career_losses.append(0)
        career_ties.append(0)

conn.close()

# Add new columns
df['Career_Wins'] = career_wins
df['Career_Losses'] = career_losses
df['Career_Ties'] = career_ties
df['Win_Pct'] = win_pcts

# Reorder columns - add Win_Pct and record after PA/HR
new_order = [
    'PID',
    'FirstName',
    'LastName',
    'Position',
    'Age',
    'PA',
    'HR',
    'Win_Pct',
    'Career_Wins',
    'Career_Losses',
    'Career_Ties',
    'convBA_F25',
    'convBA_S25',
    'convBA_W25',
    'convBA+_F25',
    'convBA+_S25',
    'convBA+_W25',
    'Weighted_convBA+',
    'Def_Spectrum_Pts',
    'Age_Factor',
    'Manual_Adj',
    'Ranking_Score',
    'Availability'
]

df = df[new_order]

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns (M, N, O - adjusted for new columns)
for col_idx in [12, 13, 14]:  # convBA_F25, S25, W25
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column (H)
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=8)  # Win_Pct column
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Added columns: Win_Pct, Career_Wins, Career_Losses, Career_Ties")
print(f"  Backup: {backup_file}")

print(f"\nTop 10 by Ranking_Score with Win%:")
print(df[['FirstName', 'LastName', 'Win_Pct', 'Career_Wins', 'Career_Losses', 'Ranking_Score']].head(10).to_string(index=False))

