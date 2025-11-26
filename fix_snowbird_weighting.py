#!/usr/bin/env python3
"""
Fix weighted convBA+ to properly handle snowbirds
- Use only seasons they actually played
- Normalize weights so they add to 100%
- F25: 50%, S25: 30%, W25: 20% (but only for seasons with data)

Example: If only played W25, use 100% of W25
         If played S25 + W25, use 60% S25 + 40% W25 (30/(30+20) and 20/(30+20))
"""

import pandas as pd
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

print("\nRecalculating weighted convBA+ (normalized for seasons played)...")

# Recalculate weighted convBA+ with normalized weights
weighted_convba_plus = []
for idx, row in df.iterrows():
    f25 = row['convBA+_F25'] if pd.notna(row['convBA+_F25']) else None
    s25 = row['convBA+_S25'] if pd.notna(row['convBA+_S25']) else None
    w25 = row['convBA+_W25'] if pd.notna(row['convBA+_W25']) else None
    
    # Base weights (F25: 50%, S25: 30%, W25: 20%)
    weights = {
        'f25': 0.50,
        's25': 0.30,
        'w25': 0.20
    }
    
    # Calculate weighted average using only seasons played (ignore 0 values = didn't play)
    seasons = {}
    if f25 is not None and f25 > 0:
        seasons['f25'] = (f25, weights['f25'])
    if s25 is not None and s25 > 0:
        seasons['s25'] = (s25, weights['s25'])
    if w25 is not None and w25 > 0:
        seasons['w25'] = (w25, weights['w25'])
    
    if seasons:
        # Normalize weights to sum to 1.0
        total_weight = sum(weight for _, weight in seasons.values())
        normalized_seasons = {k: (value, weight/total_weight) for k, (value, weight) in seasons.items()}
        
        # Calculate weighted average
        weighted_avg = sum(value * norm_weight for value, norm_weight in normalized_seasons.values())
        weighted_convba_plus.append(round(weighted_avg, 1))
    else:
        weighted_convba_plus.append(None)

df['Weighted_convBA+'] = weighted_convba_plus

# Recalculate Ranking_Score
print("Recalculating Ranking_Score...")
ranking_scores = []
for idx, row in df.iterrows():
    base = row['Weighted_convBA+'] if pd.notna(row['Weighted_convBA+']) else 0
    defense = row['Def_Spectrum_Pts']
    age = row['Age_Factor']
    manual = row['Manual_Adj']
    avail_adj = row['Availability_Adj']
    
    if base > 0:  # Only rank players with stats
        score = round(base + defense + age + manual + avail_adj, 1)
        ranking_scores.append(score)
    else:
        ranking_scores.append(None)

df['Ranking_Score'] = ranking_scores

# Sort by Ranking_Score (descending)
df = df.sort_values('Ranking_Score', ascending=False, na_position='last')

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns (N, O, P)
for col_idx in [14, 15, 16]:  # convBA_F25, S25, W25
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column (K)
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)  # Win_Pct column
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Fixed: Weighted_convBA+ now normalized for seasons played")
print(f"  Snowbirds (W25 only) now use 100% of their W25 performance")
print(f"  Recalculated Ranking_Score")
print(f"  Re-sorted by Ranking_Score")
print(f"  Backup: {backup_file}")

print(f"\nTop 20 rankings:")
display_cols = ['FirstName', 'LastName', 'Position', 'convBA+_F25', 'convBA+_S25', 'convBA+_W25', 'Weighted_convBA+', 'Ranking_Score']
print(df[display_cols].head(20).to_string(index=False))

print(f"\nSnowbirds who moved up (W25 only players):")
snowbirds = df[(df['convBA+_F25'].isna()) & (df['convBA+_S25'].isna()) & (df['convBA+_W25'].notna())]
if len(snowbirds) > 0:
    print(f"  Found {len(snowbirds)} W25-only players")
    print(snowbirds[['FirstName', 'LastName', 'Position', 'convBA+_W25', 'Weighted_convBA+', 'Ranking_Score']].head(10).to_string(index=False))

