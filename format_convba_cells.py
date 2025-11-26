#!/usr/bin/env python3
"""
Format convBA columns in Excel to display as .XXX (e.g., .230, .406)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
from datetime import datetime
import shutil

EXCEL_FILE = 'w26rankings.xlsx'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

print(f"Loading {EXCEL_FILE}...")
# Load the workbook
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Find the convBA columns (G, H, I for convBA_F25, convBA_S25, convBA_W25)
print("Applying .000 format to convBA columns...")

# Get header row to find column indices
headers = []
for cell in ws[1]:
    headers.append(cell.value)

# Find convBA column indices
convba_cols = []
for idx, header in enumerate(headers):
    if header and 'convBA_' in header and '+' not in header:
        convba_cols.append(idx + 1)  # Excel columns are 1-indexed
        print(f"  Found {header} in column {idx + 1}")

# Apply custom number format to convBA columns
# Format code: ".000" will display as .230, .406, etc.
for col_idx in convba_cols:
    for row in range(2, ws.max_row + 1):  # Skip header row
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

print(f"\nSaving {EXCEL_FILE}...")
wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Applied .000 format to convBA columns")
print(f"  Values will display as .230, .406, etc.")
print(f"  Backup: {backup_file}")

