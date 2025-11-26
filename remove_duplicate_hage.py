#!/usr/bin/env python3
"""
Remove duplicate Mike Hage entry (the one with no position)
"""

import pandas as pd
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# Find Mike Hage entries
hages = df[df['LastName'] == 'Hage']
print(f"\nFound {len(hages)} Mike Hage entries:")
print(hages[['PID', 'FirstName', 'LastName', 'Position', 'PA', 'Ranking_Score']].to_string(index=False))

# Remove the one with no position
before_count = len(df)
df = df[~((df['LastName'] == 'Hage') & (df['Position'].isna()))]
after_count = len(df)

removed = before_count - after_count
print(f"\nRemoved {removed} duplicate entry")

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns
for col_idx in [14, 15, 16]:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Removed Mike Hage duplicate (no position)")
print(f"  Remaining players: {len(df)}")
print(f"  Backup: {backup_file}")

# Verify
remaining_hages = df[df['LastName'] == 'Hage']
print(f"\nRemaining Mike Hage entries: {len(remaining_hages)}")
if len(remaining_hages) > 0:
    print(remaining_hages[['PID', 'FirstName', 'LastName', 'Position', 'PA', 'Ranking_Score']].to_string(index=False))

