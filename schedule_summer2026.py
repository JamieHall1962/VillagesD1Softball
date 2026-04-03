#!/usr/bin/env python3
"""
Villages D1 Softball - Summer 2026 Schedule Generator
8 Teams • Triple Round Robin • 21 Game Days • 84 Total Games
All games at Everglades Fields 1 & 4
"""

import datetime
import csv
import random
from collections import defaultdict
import copy

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Team names (update after draft) ──────────────────────────────
TEAMS = [
    'Stars', 'Raptors', 'Lightning Strikes', 'Shorebirds',
    'Bad News Bears', 'Rebels', 'Xtreme', 'Thunder'
]

FIELDS = ['Field 1', 'Field 4']
TIMES = ['9:00 AM', '10:30 AM']
LOCATION = 'Everglades'

# ── Key dates ────────────────────────────────────────────────────
PRACTICE_DATE = datetime.date(2026, 5, 6)       # Wednesday
FIRST_GAME = datetime.date(2026, 5, 8)          # Friday
RWB_TOURNAMENT = datetime.date(2026, 7, 3)      # Friday - Red White & Blue
LAST_GAME = datetime.date(2026, 7, 22)          # Wednesday
QUARTERFINALS = datetime.date(2026, 7, 24)      # Friday
SEMIFINALS = datetime.date(2026, 7, 29)         # Wednesday
CHAMPIONSHIP = datetime.date(2026, 7, 31)       # Friday


def generate_play_dates():
    """Generate the 21 game dates: every Wed & Fri from May 8 to July 22, skipping July 3"""
    play_dates = []
    current = FIRST_GAME

    while current <= LAST_GAME:
        if current.weekday() in [2, 4]:  # Wed=2, Fri=4
            if current != RWB_TOURNAMENT:
                play_dates.append(current)
        current += datetime.timedelta(days=1)

    return play_dates


def generate_triple_round_robin():
    """
    Generate a triple round-robin for 8 teams using the circle method.
    3 complete round-robins x 7 rounds x 4 games = 84 games across 21 dates.
    Home/away balanced: each team gets 10 or 11 home games, each pair splits 2-1.
    """
    n = len(TEAMS)
    game_dates = generate_play_dates()

    print(f"Game dates: {len(game_dates)} (need 21)")
    assert len(game_dates) == 21, f"Expected 21 game dates, got {len(game_dates)}"

    # Step 1: Generate the 3 round-robin pairings using circle method
    all_rounds = []
    for rr in range(3):
        teams_idx = list(range(n))
        for round_num in range(n - 1):
            round_pairs = []
            for i in range(n // 2):
                if i == 0:
                    idx_a = teams_idx[0]
                    idx_b = teams_idx[-1]
                else:
                    idx_a = teams_idx[i]
                    idx_b = teams_idx[n - 1 - i]
                round_pairs.append((idx_a, idx_b))
            all_rounds.append((rr, round_num, round_pairs))
            teams_idx = [teams_idx[0]] + [teams_idx[-1]] + teams_idx[1:-1]

    # Step 2: Assign home/away optimally
    # For each pair: first meeting one hosts, second the other, third goes to lesser total
    pair_meetings = defaultdict(int)
    home_counts = {i: 0 for i in range(n)}

    schedule = []
    for rr, round_num, round_pairs in all_rounds:
        date_index = rr * 7 + round_num
        round_games = []

        for idx_a, idx_b in round_pairs:
            pair = tuple(sorted([idx_a, idx_b]))
            meeting = pair_meetings[pair]
            pair_meetings[pair] += 1

            if meeting == 0:
                home_idx, away_idx = idx_a, idx_b
            elif meeting == 1:
                home_idx, away_idx = idx_b, idx_a
            else:
                if home_counts[idx_a] <= home_counts[idx_b]:
                    home_idx, away_idx = idx_a, idx_b
                else:
                    home_idx, away_idx = idx_b, idx_a

            home_counts[home_idx] += 1

            round_games.append({
                'date': game_dates[date_index],
                'round': date_index + 1,
                'rr': rr + 1,
                'home': TEAMS[home_idx],
                'away': TEAMS[away_idx],
                'field': '',
                'time': ''
            })

        schedule.extend(round_games)

    return schedule


def assign_balanced_venues(schedule):
    """
    Assign venues balancing field and time distribution across teams.
    Each round has 4 games → 2 at 9:00 (Fields 1,4) and 2 at 10:30 (Fields 1,4).
    """
    team_field_count = {team: {f: 0 for f in FIELDS} for team in TEAMS}
    team_time_count = {team: {t: 0 for t in TIMES} for team in TEAMS}

    rounds = defaultdict(list)
    for game in schedule:
        rounds[game['round']].append(game)

    venues = [
        ('Field 1', '9:00 AM'),
        ('Field 4', '9:00 AM'),
        ('Field 1', '10:30 AM'),
        ('Field 4', '10:30 AM'),
    ]

    for round_num in sorted(rounds.keys()):
        round_games = rounds[round_num]
        assert len(round_games) == 4, f"Round {round_num} has {len(round_games)} games"

        used_venues = set()

        for game in round_games:
            best_venue = None
            best_score = float('inf')

            for venue in venues:
                if venue in used_venues:
                    continue
                field, time = venue
                score = (team_field_count[game['home']][field] +
                         team_field_count[game['away']][field] +
                         team_time_count[game['home']][time] +
                         team_time_count[game['away']][time])
                if score < best_score:
                    best_score = score
                    best_venue = venue

            if best_venue:
                field, time = best_venue
                game['field'] = field
                game['time'] = time
                used_venues.add(best_venue)
                team_field_count[game['home']][field] += 1
                team_field_count[game['away']][field] += 1
                team_time_count[game['home']][time] += 1
                team_time_count[game['away']][time] += 1

    return schedule


def can_swap_games(game1, game2, schedule):
    """Check if two games can swap rounds without creating conflicts"""
    if game1['round'] == game2['round']:
        return False

    round1_teams = set()
    round2_teams = set()

    for game in schedule:
        if game['round'] == game1['round'] and game is not game1:
            round1_teams.update([game['home'], game['away']])
        elif game['round'] == game2['round'] and game is not game2:
            round2_teams.update([game['home'], game['away']])

    if {game2['home'], game2['away']} & round1_teams:
        return False
    if {game1['home'], game1['away']} & round2_teams:
        return False

    return True


def swap_games(game1, game2):
    """Swap round and date between two games"""
    game1['round'], game2['round'] = game2['round'], game1['round']
    game1['date'], game2['date'] = game2['date'], game1['date']


def score_consecutive(schedule):
    """Score consecutive home or away games (lower is better)"""
    sorted_sched = sorted(schedule, key=lambda x: x['date'])
    team_seq = {team: [] for team in TEAMS}

    for game in sorted_sched:
        team_seq[game['home']].append('H')
        team_seq[game['away']].append('A')

    total = 0
    details = {}
    for team in TEAMS:
        seq = team_seq[team]
        max_h = max_a = cur_h = cur_a = 0
        for g in seq:
            if g == 'H':
                cur_h += 1; cur_a = 0; max_h = max(max_h, cur_h)
            else:
                cur_a += 1; cur_h = 0; max_a = max(max_a, cur_a)
        violations = max(0, max_h - 3) + max(0, max_a - 3)
        total += violations
        details[team] = {'max_home': max_h, 'max_away': max_a, 'violations': violations}

    return total, details


def optimize_schedule(schedule, max_iterations=15000):
    """Optimize to minimize consecutive home/away streaks"""
    print("Optimizing consecutive home/away balance...")
    schedule = copy.deepcopy(schedule)
    best_score, best_details = score_consecutive(schedule)
    best_schedule = copy.deepcopy(schedule)
    improvements = 0

    print(f"  Initial score: {best_score}")

    for iteration in range(max_iterations):
        current_score, current_details = score_consecutive(schedule)

        problem_teams = [(t, d) for t, d in current_details.items() if d['violations'] > 0]
        problem_teams.sort(key=lambda x: x[1]['violations'], reverse=True)

        if not problem_teams:
            print("  No consecutive violations remaining!")
            break

        worst_team = problem_teams[0][0]
        worst_games = [g for g in schedule if g['home'] == worst_team or g['away'] == worst_team]

        swap_made = False
        for g1 in worst_games:
            if swap_made:
                break
            for g2 in schedule:
                if g2['round'] != g1['round'] and can_swap_games(g1, g2, schedule):
                    swap_games(g1, g2)
                    new_score, _ = score_consecutive(schedule)
                    if new_score < current_score:
                        if new_score < best_score:
                            best_score = new_score
                            best_schedule = copy.deepcopy(schedule)
                            improvements += 1
                        swap_made = True
                        break
                    else:
                        swap_games(g1, g2)

        if not swap_made:
            g1, g2 = random.choice(schedule), random.choice(schedule)
            if can_swap_games(g1, g2, schedule):
                swap_games(g1, g2)
                new_score, _ = score_consecutive(schedule)
                if new_score < best_score:
                    best_score = new_score
                    best_schedule = copy.deepcopy(schedule)
                    improvements += 1
                else:
                    swap_games(g1, g2)

    print(f"  Final score: {best_score} ({improvements} improvements)")

    best_schedule = assign_balanced_venues(best_schedule)
    _, best_details = score_consecutive(best_schedule)
    return best_schedule, best_details


def verify_schedule(schedule):
    """Verify all constraints"""
    errors = []
    n_teams = len(TEAMS)
    expected_games = n_teams * (n_teams - 1) // 2 * 3  # triple round robin

    if len(schedule) != expected_games:
        errors.append(f"Wrong game count: {len(schedule)} (expected {expected_games})")

    team_counts = {t: 0 for t in TEAMS}
    for g in schedule:
        team_counts[g['home']] += 1
        team_counts[g['away']] += 1
    for t, c in team_counts.items():
        if c != 21:
            errors.append(f"{t} plays {c} games (expected 21)")

    matchup_counts = defaultdict(int)
    for g in schedule:
        pair = tuple(sorted([g['home'], g['away']]))
        matchup_counts[pair] += 1
    for pair, count in matchup_counts.items():
        if count != 3:
            errors.append(f"{pair[0]} vs {pair[1]}: {count} games (expected 3)")

    rounds = defaultdict(list)
    for g in schedule:
        rounds[g['round']].append(g)
    for rnd, games in rounds.items():
        teams_in_round = []
        for g in games:
            teams_in_round.extend([g['home'], g['away']])
        for t in TEAMS:
            if teams_in_round.count(t) > 1:
                errors.append(f"Round {rnd}: {t} appears {teams_in_round.count(t)} times")

    return errors


def print_schedule(schedule):
    """Print the full schedule"""
    print(f"\n{'='*70}")
    print("Division One - Summer 2026 Schedule")
    print(f"All games played at {LOCATION}")
    print(f"Revised {datetime.date.today().strftime('%B %d, %Y')}")
    print(f"{'='*70}\n")

    # Practice day
    print(f"{PRACTICE_DATE.strftime('%A, %B %d, %Y')}")
    print("  PRACTICE DAY - Team Preparation")
    print()

    games_by_date = defaultdict(list)
    for g in schedule:
        games_by_date[g['date']].append(g)

    for date in sorted(games_by_date.keys()):
        games = sorted(games_by_date[date],
                       key=lambda g: (0 if '9:00' in g['time'] else 1, g['field']))
        print(f"{date.strftime('%A, %B %d, %Y')}")
        for g in games:
            time_str = g['time'].replace(' AM', '').replace(' PM', '')
            field_num = g['field'].split()[-1]
            print(f"  {time_str:<8}{g['away']} at {g['home']} on {field_num}")
        print()

    # Special events
    print(f"{RWB_TOURNAMENT.strftime('%A, %B %d, %Y')}")
    print("  RED WHITE & BLUE TOURNAMENT")
    print()

    print(f"{'-'*70}")
    print("PLAYOFFS")
    print(f"{'-'*70}")
    print(f"{QUARTERFINALS.strftime('%A, %B %d, %Y')}")
    print("  Quarterfinals")
    print(f"{SEMIFINALS.strftime('%A, %B %d, %Y')}")
    print("  Semifinals")
    print(f"{CHAMPIONSHIP.strftime('%A, %B %d, %Y')}")
    print("  Championship")


def analyze_schedule(schedule, details):
    """Print schedule analysis"""
    print(f"\n{'='*70}")
    print("SCHEDULE ANALYSIS")
    print(f"{'='*70}")

    # Home/Away balance
    home_counts = {t: 0 for t in TEAMS}
    away_counts = {t: 0 for t in TEAMS}
    for g in schedule:
        home_counts[g['home']] += 1
        away_counts[g['away']] += 1

    print(f"\n{'Team':<20} {'Home':<6} {'Away':<6} {'Total':<6}")
    print("-" * 40)
    for t in TEAMS:
        print(f"{t:<20} {home_counts[t]:<6} {away_counts[t]:<6} {home_counts[t]+away_counts[t]:<6}")

    # Field balance
    team_field = {t: {f: 0 for f in FIELDS} for t in TEAMS}
    team_time = {t: {tm: 0 for tm in TIMES} for t in TEAMS}
    for g in schedule:
        for t in [g['home'], g['away']]:
            team_field[t][g['field']] += 1
            team_time[t][g['time']] += 1

    print(f"\n{'Team':<20} {'Field 1':<10} {'Field 4':<10}")
    print("-" * 40)
    for t in TEAMS:
        print(f"{t:<20} {team_field[t]['Field 1']:<10} {team_field[t]['Field 4']:<10}")

    print(f"\n{'Team':<20} {'9:00 AM':<10} {'10:30 AM':<10}")
    print("-" * 40)
    for t in TEAMS:
        print(f"{t:<20} {team_time[t]['9:00 AM']:<10} {team_time[t]['10:30 AM']:<10}")

    # Consecutive streaks
    print(f"\n{'Team':<20} {'Max Home':<10} {'Max Away':<10} {'Status'}")
    print("-" * 50)
    for t in TEAMS:
        d = details[t]
        status = "OK" if d['violations'] == 0 else f"STREAK ({d['violations']})"
        print(f"{t:<20} {d['max_home']:<10} {d['max_away']:<10} {status}")

    # Matchup matrix
    print(f"\n{'='*70}")
    print("MATCHUP MATRIX (home games listed)")
    print(f"{'='*70}")
    host_counts = defaultdict(lambda: defaultdict(int))
    for g in schedule:
        host_counts[g['home']][g['away']] += 1

    header = f"{'':>20}"
    for t in TEAMS:
        header += f" {t[:6]:>6}"
    print(header)

    for home in TEAMS:
        row = f"{home:<20}"
        for away in TEAMS:
            if home == away:
                row += f" {'--':>6}"
            else:
                row += f" {host_counts[home][away]:>6}"
        print(row)


def export_to_excel(schedule):
    """Export schedule to a clean Excel file"""
    if not EXCEL_AVAILABLE:
        print("openpyxl not installed. Install with: pip install openpyxl")
        return

    wb = Workbook()
    wb.remove(wb.active)

    # ── Main schedule sheet ──────────────────────────────────────
    ws = wb.create_sheet("Summer 2026 Schedule")
    ws.sheet_view.showGridLines = False

    col_widths = {'A': 4, 'B': 44, 'C': 4, 'D': 44, 'E': 4}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font = Font(size=18, bold=True, color="FFFFFF")
    subtitle_font = Font(size=12, italic=True, color="FFFFFF")
    date_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    date_font = Font(size=12, bold=True, color="1F4E79")
    game_font = Font(size=10, color="333333")
    special_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    special_font = Font(size=10, italic=True, bold=True, color="C55A11")
    playoff_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    playoff_font = Font(size=11, bold=True, color="375623")

    ws.merge_cells('B2:D2')
    ws['B2'] = "Division One - Summer 2026"
    ws['B2'].font = title_font
    ws['B2'].fill = title_fill
    ws['B2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B3:D3')
    ws['B3'] = f"All games played at {LOCATION} • Fields 1 & 4"
    ws['B3'].font = subtitle_font
    ws['B3'].fill = title_fill
    ws['B3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B4:D4')
    ws['B4'] = f"Revised {datetime.date.today().strftime('%B %d, %Y')}"
    ws['B4'].font = Font(size=10, italic=True, color="666666")
    ws['B4'].alignment = Alignment(horizontal='center')

    games_by_date = defaultdict(list)
    for g in schedule:
        games_by_date[g['date']].append(g)

    all_dates = sorted(games_by_date.keys())
    all_dates_with_special = []

    all_dates_with_special.append((PRACTICE_DATE, 'practice'))
    for d in all_dates:
        all_dates_with_special.append((d, 'game'))
        if d < RWB_TOURNAMENT:
            next_dates = [x for x in all_dates if x > d]
            if next_dates and next_dates[0] > RWB_TOURNAMENT:
                all_dates_with_special.append((RWB_TOURNAMENT, 'rwb'))

    if RWB_TOURNAMENT not in [x[0] for x in all_dates_with_special]:
        for i, (d, t) in enumerate(all_dates_with_special):
            if d > RWB_TOURNAMENT:
                all_dates_with_special.insert(i, (RWB_TOURNAMENT, 'rwb'))
                break

    # Group into Wed/Fri pairs for side-by-side layout
    wednesdays = [(d, t) for d, t in all_dates_with_special if d.weekday() == 2]
    fridays = [(d, t) for d, t in all_dates_with_special if d.weekday() == 4]

    current_row = 6

    max_pairs = max(len(wednesdays), len(fridays))
    for i in range(max_pairs):
        start_row = current_row

        # Wednesday column (B)
        if i < len(wednesdays):
            wed_date, wed_type = wednesdays[i]
            ws[f'B{current_row}'] = wed_date.strftime('%A, %B %d')
            ws[f'B{current_row}'].font = date_font
            ws[f'B{current_row}'].fill = date_fill
            current_row += 1

            if wed_type == 'practice':
                ws[f'B{current_row}'] = "PRACTICE DAY - Team Preparation"
                ws[f'B{current_row}'].font = special_font
                ws[f'B{current_row}'].fill = special_fill
                current_row += 1
            elif wed_type == 'game' and wed_date in games_by_date:
                games = sorted(games_by_date[wed_date],
                               key=lambda g: (0 if '9:00' in g['time'] else 1, g['field']))
                for g in games:
                    time_str = g['time'].replace(' AM', '').replace(' PM', '')
                    field_num = g['field'].split()[-1]
                    ws[f'B{current_row}'] = f"{time_str}  {g['away']} at {g['home']} on {field_num}"
                    ws[f'B{current_row}'].font = game_font
                    current_row += 1

        # Friday column (D)
        fri_row = start_row
        if i < len(fridays):
            fri_date, fri_type = fridays[i]
            ws[f'D{fri_row}'] = fri_date.strftime('%A, %B %d')
            ws[f'D{fri_row}'].font = date_font
            ws[f'D{fri_row}'].fill = date_fill
            fri_row += 1

            if fri_type == 'rwb':
                ws[f'D{fri_row}'] = "RED WHITE & BLUE TOURNAMENT"
                ws[f'D{fri_row}'].font = special_font
                ws[f'D{fri_row}'].fill = special_fill
                fri_row += 1
            elif fri_type == 'game' and fri_date in games_by_date:
                games = sorted(games_by_date[fri_date],
                               key=lambda g: (0 if '9:00' in g['time'] else 1, g['field']))
                for g in games:
                    time_str = g['time'].replace(' AM', '').replace(' PM', '')
                    field_num = g['field'].split()[-1]
                    ws[f'D{fri_row}'] = f"{time_str}  {g['away']} at {g['home']} on {field_num}"
                    ws[f'D{fri_row}'].font = game_font
                    fri_row += 1

        current_row = max(current_row, fri_row) + 1

    # Playoffs section
    current_row += 1
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = "PLAYOFFS"
    ws[f'B{current_row}'].font = Font(size=14, bold=True, color="375623")
    ws[f'B{current_row}'].fill = playoff_fill
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1

    for label, date in [("Quarterfinals", QUARTERFINALS),
                         ("Semifinals", SEMIFINALS),
                         ("Championship", CHAMPIONSHIP)]:
        ws[f'B{current_row}'] = date.strftime('%A, %B %d')
        ws[f'B{current_row}'].font = date_font
        ws[f'B{current_row}'].fill = playoff_fill
        ws[f'D{current_row}'] = label
        ws[f'D{current_row}'].font = playoff_font
        ws[f'D{current_row}'].fill = playoff_fill
        current_row += 1

    # ── Grid sheet ───────────────────────────────────────────────
    ws_grid = wb.create_sheet("Schedule Grid")
    ws_grid.sheet_view.showGridLines = False

    ws_grid.merge_cells('A1:M1')
    ws_grid['A1'] = "Division One - Summer 2026 Schedule"
    ws_grid['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws_grid['A1'].alignment = Alignment(horizontal='center')

    ws_grid.merge_cells('A2:M2')
    ws_grid['A2'] = f"All games at {LOCATION} • Fields 1 & 4"
    ws_grid['A2'].font = Font(size=11, italic=True, color="666666")
    ws_grid['A2'].alignment = Alignment(horizontal='center')

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    headers = ['Date', 'Day', 'Rnd',
               '9:00 Field', '9:00 Away', '9:00 Home',
               '9:00 Field', '9:00 Away', '9:00 Home',
               '10:30 Field', '10:30 Away', '10:30 Home',
               '10:30 Field', '10:30 Away', '10:30 Home']

    for col_idx, header in enumerate(headers, 1):
        cell = ws_grid.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    ws_grid.column_dimensions['A'].width = 12
    ws_grid.column_dimensions['B'].width = 6
    ws_grid.column_dimensions['C'].width = 5
    for c in range(4, 16):
        ws_grid.column_dimensions[get_column_letter(c)].width = 14

    row = 5
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    for date in sorted(games_by_date.keys()):
        games = sorted(games_by_date[date],
                       key=lambda g: (0 if '9:00' in g['time'] else 1, g['field']))

        ws_grid.cell(row=row, column=1, value=date.strftime('%m/%d/%Y'))
        ws_grid.cell(row=row, column=2, value=date.strftime('%a'))
        ws_grid.cell(row=row, column=3, value=games[0]['round'])

        col = 4
        for g in games:
            field_num = g['field'].split()[-1]
            ws_grid.cell(row=row, column=col, value=field_num)
            ws_grid.cell(row=row, column=col + 1, value=g['away'])
            ws_grid.cell(row=row, column=col + 2, value=g['home'])
            col += 3

        if row % 2 == 0:
            for c in range(1, 16):
                ws_grid.cell(row=row, column=c).fill = alt_fill

        row += 1

    # RWB tournament row
    ws_grid.cell(row=row, column=1, value=RWB_TOURNAMENT.strftime('%m/%d/%Y'))
    ws_grid.cell(row=row, column=2, value='Fri')
    ws_grid.merge_cells(f'D{row}:O{row}')
    ws_grid.cell(row=row, column=4, value="RED WHITE & BLUE TOURNAMENT - NO REGULAR GAMES")
    ws_grid.cell(row=row, column=4).font = Font(italic=True, color="C55A11")
    ws_grid.cell(row=row, column=4).alignment = Alignment(horizontal='center')

    filename = "villages_d1_summer_2026_schedule.xlsx"
    wb.save(filename)
    print(f"Excel file created: {filename}")


def export_csv(schedule):
    """Export to CSV for data entry"""
    filename = "villages_d1_summer_2026_schedule.csv"

    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Date', 'Day', 'Round', 'Time', 'Field', 'Away Team', 'Home Team'])

        for g in sorted(schedule, key=lambda x: (x['date'], x['time'], x['field'])):
            writer.writerow([
                g['date'].strftime('%m/%d/%Y'),
                g['date'].strftime('%a'),
                g['round'],
                g['time'],
                g['field'],
                g['away'],
                g['home']
            ])

    print(f"CSV file created: {filename}")


def main():
    print("Villages D1 Softball - Summer 2026 Schedule Generator")
    print("8 Teams • Triple Round Robin • 21 Game Days • 84 Games")
    print(f"Location: {LOCATION} (Fields 1 & 4)")
    print("=" * 70)

    game_dates = generate_play_dates()
    print(f"\nPractice day: {PRACTICE_DATE.strftime('%A, %B %d, %Y')}")
    print(f"First game:   {FIRST_GAME.strftime('%A, %B %d, %Y')}")
    print(f"Last game:    {LAST_GAME.strftime('%A, %B %d, %Y')}")
    print(f"Game dates:   {len(game_dates)}")
    print(f"\nAll game dates:")
    for i, d in enumerate(game_dates, 1):
        print(f"  {i:2d}. {d.strftime('%A, %B %d, %Y')}")

    # Generate base schedule
    print(f"\nGenerating triple round-robin...")
    base_schedule = generate_triple_round_robin()

    # Assign initial venues
    base_schedule = assign_balanced_venues(base_schedule)

    # Verify
    errors = verify_schedule(base_schedule)
    if errors:
        print("BASE SCHEDULE ERRORS:")
        for e in errors:
            print(f"  - {e}")
        return
    print("Base schedule verified - all constraints met")

    # Optimize
    optimized, details = optimize_schedule(base_schedule, max_iterations=15000)

    # Final verify
    errors = verify_schedule(optimized)
    if errors:
        print("OPTIMIZATION BROKE CONSTRAINTS:")
        for e in errors:
            print(f"  - {e}")
        return
    print("Optimized schedule verified")

    # Output
    print_schedule(optimized)
    analyze_schedule(optimized, details)
    export_to_excel(optimized)
    export_csv(optimized)

    print(f"\nFiles created:")
    print(f"  villages_d1_summer_2026_schedule.xlsx")
    print(f"  villages_d1_summer_2026_schedule.csv")


if __name__ == "__main__":
    main()
