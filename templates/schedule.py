#!/usr/bin/env python3
"""
Villages D1 Softball Schedule Generator - Deterministic + Smart Swaps
1. Generate guaranteed valid round-robin (hard constraints always met)
2. Intelligently swap games between rounds to optimize consecutive games
3. Only allow swaps that preserve all hard constraints
"""

import datetime
import csv
import random
from collections import defaultdict
import copy

# Import Excel libraries at the top
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

TEAMS = [
    'The Sandlot', 'Xtreme', 'Stars', 'Bad News Bears', 'Norsemen', 'Clippers',
    'Rebels', 'Raptors', 'Buckeyes', 'Lightning Strikes', 'Shorebirds', 'Warhawks'
]

FIELDS = ['Field 1', 'Field 2', 'Field 3']
TIMES = ['9:00 AM', '10:30 AM']

def generate_play_dates():
    """Generate play dates starting September 3 (practice), ensuring we include Oct 15 and 17"""
    play_dates = []
    current = datetime.date(2025, 9, 3)  # Start September 3, 2025 (Wednesday - practice day)
    
    # Generate enough dates to include October 15 and 17
    while len(play_dates) < 25:  # Generate extra to ensure we get to October
        if current.weekday() in [2, 4]:  # Wed/Fri
            play_dates.append(current)
        current += datetime.timedelta(days=1)
    
    return play_dates

def generate_guaranteed_valid_schedule():
    """Generate a guaranteed valid round-robin schedule"""
    print("Generating guaranteed valid base schedule...")
    
    n = len(TEAMS)
    schedule = []
    play_dates = generate_play_dates()
    
    # Remove BOTH practice day and blackout date from game scheduling
    practice_date = datetime.date(2025, 9, 3)
    blackout_date = datetime.date(2025, 10, 17)
    game_dates = [date for date in play_dates if date not in [practice_date, blackout_date]]
    
    print(f"Total play dates: {len(play_dates)}")
    print(f"Game dates (excluding practice and blackout): {len(game_dates)}")
    print(f"Need 22 rounds, have {len(game_dates)} game dates")
    
    # Standard round-robin using circle method (proven to work)
    teams = list(range(n))
    
    # We need exactly 22 rounds, so let's ensure we use exactly 22 game dates
    if len(game_dates) < 22:
        print(f"ERROR: Not enough game dates! Need 22, have {len(game_dates)}")
        return []
    
    # Take exactly the first 22 game dates for our 22 rounds
    round_dates = game_dates[:22]
    
    print("Round dates (games only, excluding practice day):")
    for i, date in enumerate(round_dates, 1):
        print(f"  Round {i:2d}: {date.strftime('%A, %B %d, %Y')}")
    
    # First round-robin (11 rounds)
    for round_num in range(n - 1):
        round_games = []
        
        for i in range(n // 2):
            if i == 0:
                home_idx = 0
                away_idx = teams[1]
            else:
                home_idx = teams[i + 1]
                away_idx = teams[n - i]
        
            round_games.append({
                'date': round_dates[round_num],
                'round': round_num + 1,
                'home': TEAMS[home_idx],
                'away': TEAMS[away_idx],
                'field': '',
                'time': ''
            })
        
        schedule.extend(round_games)
        teams = [teams[0]] + [teams[-1]] + teams[1:-1]
    
    # Second round-robin (reverse home/away)
    first_round = schedule.copy()
    
    for i, game in enumerate(first_round):
        round_number = 12 + (i // 6)
        date_index = 11 + (i // 6)  # Start from 12th date (index 11)
        
        schedule.append({
            'date': round_dates[date_index],
            'round': round_number,
            'home': game['away'],
            'away': game['home'],
            'field': '',
            'time': ''
        })
    
    # Assign balanced venues
    schedule = assign_balanced_venues(schedule)
    
    print(f"✅ Valid schedule created: {len(schedule)} games across {len(set(game['date'] for game in schedule))} dates")
    print(f"✅ Season starts on {min(game['date'] for game in schedule).strftime('%A, %B %d, %Y')} (first game day)")
    print(f"✅ Season ends on {max(game['date'] for game in schedule).strftime('%A, %B %d, %Y')}")
    print(f"✅ September 3rd reserved as practice day only")
    return schedule

def assign_balanced_venues(schedule):
    """Assign venues ensuring exactly 3 games at 9:00 and 3 at 10:30, using all fields"""
    team_field_count = {team: {field: 0 for field in FIELDS} for team in TEAMS}
    team_time_count = {team: {time: 0 for time in TIMES} for team in TEAMS}
    
    # Group by round
    rounds = defaultdict(list)
    for game in schedule:
        rounds[game['round']].append(game)
    
    # Define venue structure: exactly 6 games per round
    # 3 games at 9:00 AM (Fields 1, 2, 3) and 3 games at 10:30 AM (Fields 1, 2, 3)
    venues = [
        ('Field 1', '9:00 AM'),
        ('Field 2', '9:00 AM'), 
        ('Field 3', '9:00 AM'),
        ('Field 1', '10:30 AM'),
        ('Field 2', '10:30 AM'),
        ('Field 3', '10:30 AM')
    ]
    
    for round_num in sorted(rounds.keys()):
        round_games = rounds[round_num]
        
        # Ensure we have exactly 6 games
        if len(round_games) != 6:
            print(f"Warning: Round {round_num} has {len(round_games)} games instead of 6")
            continue
        
        # Assign venues using the simple balanced approach that worked before
        used_venues = set()
        
        for game in round_games:
            best_venue = None
            best_score = float('inf')
            
            # Try each available venue
            for venue in venues:
                if venue in used_venues:
                    continue
                
                field, time = venue
                # Calculate balance score - prefer venues where teams have played less
                score = (team_field_count[game['home']][field] + 
                        team_field_count[game['away']][field] +
                        team_time_count[game['home']][time] + 
                        team_time_count[game['away']][time])
                
                if score < best_score:
                    best_score = score
                    best_venue = venue
            
            # Assign the best venue
            if best_venue:
                field, time = best_venue
                game['field'] = field
                game['time'] = time
                used_venues.add(best_venue)
                
                # Update counters
                team_field_count[game['home']][field] += 1
                team_field_count[game['away']][field] += 1
                team_time_count[game['home']][time] += 1
                team_time_count[game['away']][time] += 1
    
    return schedule

def can_swap_games_safely(game1, game2, schedule):
    """Check if two games can be swapped between rounds without violating any constraints"""
    if game1['round'] == game2['round']:
        return False
    
    # Get teams in each round (excluding the games we're swapping)
    round1_teams = set()
    round2_teams = set()
    
    for game in schedule:
        if game['round'] == game1['round'] and game != game1:
            round1_teams.update([game['home'], game['away']])
        elif game['round'] == game2['round'] and game != game2:
            round2_teams.update([game['home'], game['away']])
    
    # Check if game2 can go in round1
    game2_teams = {game2['home'], game2['away']}
    if game2_teams & round1_teams:
        return False
    
    # Check if game1 can go in round2
    game1_teams = {game1['home'], game1['away']}
    if game1_teams & round2_teams:
        return False
    
    return True

def swap_games(game1, game2):
    """Swap the rounds and dates of two games (preserve venue assignments)"""
    temp_round = game1['round']
    temp_date = game1['date']
    
    game1['round'] = game2['round']
    game1['date'] = game2['date']
    
    game2['round'] = temp_round
    game2['date'] = temp_date

def score_consecutive_games(schedule):
    """Score consecutive game violations"""
    sorted_schedule = sorted(schedule, key=lambda x: x['date'])
    team_sequences = {team: [] for team in TEAMS}
    
    for game in sorted_schedule:
        team_sequences[game['home']].append('H')
        team_sequences[game['away']].append('A')
    
    total_violations = 0
    violation_details = {}
    
    for team in TEAMS:
        sequence = team_sequences[team]
        max_home = max_away = 0
        current_home = current_away = 0
        
        for game_type in sequence:
            if game_type == 'H':
                current_home += 1
                current_away = 0
                max_home = max(max_home, current_home)
            else:
                current_away += 1
                current_home = 0
                max_away = max(max_away, current_away)
        
        # Count violations (>4 is violation, >3 is warning)
        violations = 0
        if max_home > 4:
            violations += max_home - 4
        if max_away > 4:
            violations += max_away - 4
        
        total_violations += violations
        
        violation_details[team] = {
            'max_home': max_home,
            'max_away': max_away,
            'violations': violations
        }
    
    return total_violations, violation_details

def optimize_consecutive_games_intelligently(schedule, max_iterations=10000):
    """Intelligently optimize consecutive games through targeted swapping"""
    print(f"Optimizing consecutive games with up to {max_iterations} smart swaps...")
    
    schedule = copy.deepcopy(schedule)
    
    # Get initial score
    best_score, best_details = score_consecutive_games(schedule)
    best_schedule = copy.deepcopy(schedule)
    
    print(f"Initial consecutive violations: {best_score}")
    
    improvements = 0
    
    for iteration in range(max_iterations):
        if iteration % 1000 == 0:
            print(f"  Iteration {iteration + 1}/{max_iterations} (best score: {best_score}, improvements: {improvements})")
        
        # Find teams with worst consecutive violations
        current_score, current_details = score_consecutive_games(schedule)
        
        # Get teams with violations, sorted by severity
        problem_teams = [(team, details) for team, details in current_details.items() if details['violations'] > 0]
        problem_teams.sort(key=lambda x: x[1]['violations'], reverse=True)
        
        if not problem_teams:
            print(f"  🎉 No consecutive violations remaining!")
            break
        
        # Focus on worst team
        worst_team = problem_teams[0][0]
        
        # Find games involving worst team that might help if swapped
        worst_team_games = [game for game in schedule if game['home'] == worst_team or game['away'] == worst_team]
        
        # Try swapping each game involving worst team
        swap_made = False
        for game1 in worst_team_games:
            if swap_made:
                break
                
            # Try swapping with games from other rounds
            for game2 in schedule:
                if game2['round'] != game1['round']:
                    if can_swap_games_safely(game1, game2, schedule):
                        # Try the swap
                        swap_games(game1, game2)
                        
                        # Score the new schedule
                        new_score, new_details = score_consecutive_games(schedule)
                        
                        if new_score < current_score:
                            # Improvement! Keep this swap
                            if new_score < best_score:
                                best_score = new_score
                                best_details = new_details
                                best_schedule = copy.deepcopy(schedule)
                                improvements += 1
                                print(f"    Improvement #{improvements}: New best score {best_score}")
                            swap_made = True
                            break
                        else:
                            # No improvement, revert
                            swap_games(game1, game2)
        
        # If no targeted swap helped, try a random swap
        if not swap_made:
            game1 = random.choice(schedule)
            game2 = random.choice(schedule)
            
            if can_swap_games_safely(game1, game2, schedule):
                swap_games(game1, game2)
                
                new_score, new_details = score_consecutive_games(schedule)
                
                if new_score < best_score:
                    best_score = new_score
                    best_details = new_details
                    best_schedule = copy.deepcopy(schedule)
                    improvements += 1
                    print(f"    Random improvement #{improvements}: New best score {best_score}")
                else:
                    swap_games(game1, game2)  # Revert
    
    print(f"Optimization complete: {improvements} improvements found")
    print(f"Final consecutive violations: {best_score}")
    
    # CRITICAL: Re-assign venues after optimization to fix conflicts
    print("Re-assigning venues to fix any conflicts...")
    best_schedule = assign_balanced_venues(best_schedule)
    
    return best_schedule, best_details

def verify_schedule_validity(schedule):
    """Verify schedule meets all hard constraints"""
    violations = []
    
    # Check total games
    if len(schedule) != 132:
        violations.append(f"Wrong number of games: {len(schedule)}")
    
    # Check each team plays 22 games
    team_counts = {team: 0 for team in TEAMS}
    for game in schedule:
        team_counts[game['home']] += 1
        team_counts[game['away']] += 1
    
    for team, count in team_counts.items():
        if count != 22:
            violations.append(f"{team} plays {count} games instead of 22")
    
    # Check each team hosts every other team exactly once
    host_counts = defaultdict(lambda: defaultdict(int))
    for game in schedule:
        host_counts[game['home']][game['away']] += 1
    
    for home_team in TEAMS:
        for away_team in TEAMS:
            if home_team != away_team:
                count = host_counts[home_team][away_team]
                if count != 1:
                    violations.append(f"{home_team} hosts {away_team} {count} times instead of 1")
    
    # Check no team appears twice in same round
    rounds = defaultdict(list)
    for game in schedule:
        rounds[game['round']].append(game)
    
    for round_num, round_games in rounds.items():
        teams_in_round = []
        for game in round_games:
            teams_in_round.extend([game['home'], game['away']])
        
        for team in TEAMS:
            count = teams_in_round.count(team)
            if count > 1:
                violations.append(f"Round {round_num}: {team} appears {count} times")
    
    return violations

def analyze_final_schedule(schedule, violation_details):
    """Analyze the final schedule"""
    print(f"\n{'='*60}")
    print("CONSECUTIVE GAMES ANALYSIS")
    print(f"{'='*60}")
    print(f"{'Team':<20} {'Max Home':<10} {'Max Away':<10} {'Violations':<12} {'Status'}")
    print("-" * 70)
    
    total_violations = 0
    teams_with_violations = 0
    
    for team in TEAMS:
        details = violation_details[team]
        max_home = details['max_home']
        max_away = details['max_away']
        violations = details['violations']
        
        if violations > 0:
            status = f"VIOLATION ({violations})"
            teams_with_violations += 1
        elif max_home > 3 or max_away > 3:
            status = "WARNING (>3)"
        else:
            status = "GOOD"
        
        total_violations += violations
        
        print(f"{team:<20} {max_home:<10} {max_away:<10} {violations:<12} {status}")
    
    print(f"\nCONSECUTIVE GAMES SUMMARY:")
    print(f"  Total violations (>4 consecutive): {total_violations}")
    print(f"  Teams with violations: {teams_with_violations}")
    
    # Field and time balance
    team_field_counts = {team: {field: 0 for field in FIELDS} for team in TEAMS}
    team_time_counts = {team: {time: 0 for time in TIMES} for team in TEAMS}
    
    for game in schedule:
        team_field_counts[game['home']][game['field']] += 1
        team_field_counts[game['away']][game['field']] += 1
        team_time_counts[game['home']][game['time']] += 1
        team_time_counts[game['away']][game['time']] += 1
    
    print(f"\n{'='*60}")
    print("FIELD & TIME BALANCE")
    print(f"{'='*60}")
    
    print(f"{'Team':<20} {'Field 1':<8} {'Field 2':<8} {'Field 3':<8}")
    print("-" * 50)
    for team in TEAMS:
        counts = team_field_counts[team]
        print(f"{team:<20} {counts['Field 1']:<8} {counts['Field 2']:<8} {counts['Field 3']:<8}")
    
    print(f"\n{'Team':<20} {'9:00 AM':<10} {'10:30 AM':<10}")
    print("-" * 45)
    for team in TEAMS:
        counts = team_time_counts[team]
        print(f"{team:<20} {counts['9:00 AM']:<10} {counts['10:30 AM']:<10}")
    
    return teams_with_violations

def get_special_event(date):
    """Get special event for a date"""
    special_events = {
        datetime.date(2025, 9, 3): "PRACTICE DAY - Team Preparation",
        datetime.date(2025, 10, 17): "BLACKOUT DAY - Villages Tournament",
    }
    return special_events.get(date)

def print_professional_schedule(schedule):
    """Print schedule exactly like the example image"""
    
    # Simple header
    print("Division One - Fall 2025 Schedule")
    print("all games played at the Saddlebrook Softball Complex")
    print(f"revised {datetime.date.today().strftime('%B %d, %Y')}")
    print()
    
    # Group games by date
    games_by_date = defaultdict(list)
    for game in schedule:
        games_by_date[game['date']].append(game)
    
    # Get ALL unique dates from the actual schedule
    all_schedule_dates = sorted(set(game['date'] for game in schedule))
    
    # Add practice day and blackout day to the display
    practice_date = datetime.date(2025, 9, 3)
    blackout_date = datetime.date(2025, 10, 17)
    
    # Combine all dates for display
    display_dates = list(set(all_schedule_dates + [practice_date, blackout_date]))
    display_dates.sort()
    
    # Separate into Wednesdays and Fridays
    wednesdays = [date for date in display_dates if date.weekday() == 2]
    fridays = [date for date in display_dates if date.weekday() == 4]
    
    print(f"DEBUG: {len(wednesdays)} Wednesdays, {len(fridays)} Fridays")
    
    # Print side by side: Wednesdays left, Fridays right
    max_dates = max(len(wednesdays), len(fridays))
    
    for i in range(max_dates):
        print(f"DEBUG: Processing pair {i+1}")
        
        # Left column (Wednesday)
        left_lines = []
        if i < len(wednesdays):
            wed_date = wednesdays[i]
            wed_games = games_by_date.get(wed_date, [])
            wed_special = get_special_event(wed_date)
            
            print(f"DEBUG: Wednesday {wed_date} has {len(wed_games)} games, special: {wed_special}")
            
            # Date header
            left_lines.append(f"{wed_date.strftime('%A, %B %d, %Y')}")
            
            if wed_special:
                left_lines.append(wed_special)
            elif wed_games:
                # Sort games by time, then field
                sorted_games = sorted(wed_games, key=lambda g: (0 if '9:00' in g['time'] else 1, int(g['field'].split()[-1])))
                
                for game in sorted_games:
                    time_str = game['time'].replace(' AM', '').replace(' PM', '')
                    field_num = game['field'].split()[-1]
                    game_line = f"{time_str:<8}{game['away']} at {game['home']} on {field_num}"
                    left_lines.append(game_line)
            else:
                left_lines.append("No games scheduled")
        
        # Right column (Friday)
        right_lines = []
        if i < len(fridays):
            fri_date = fridays[i]
            fri_games = games_by_date.get(fri_date, [])
            fri_special = get_special_event(fri_date)
            
            print(f"DEBUG: Friday {fri_date} has {len(fri_games)} games, special: {fri_special}")
            
            # Date header
            right_lines.append(f"{fri_date.strftime('%A, %B %d, %Y')}")
            
            if fri_special:
                right_lines.append(fri_special)
            elif fri_games:
                # Sort games by time, then field
                sorted_games = sorted(fri_games, key=lambda g: (0 if '9:00' in g['time'] else 1, int(g['field'].split()[-1])))
                
                for game in sorted_games:
                    time_str = game['time'].replace(' AM', '').replace(' PM', '')
                    field_num = game['field'].split()[-1]
                    game_line = f"{time_str:<8}{game['away']} at {game['home']} on {field_num}"
                    right_lines.append(game_line)
            else:
                right_lines.append("No games scheduled")
        
        print(f"DEBUG: Left lines: {len(left_lines)}, Right lines: {len(right_lines)}")
        
        # Pad shorter column with blank lines to make them equal length
        max_lines = max(len(left_lines), len(right_lines))
        while len(left_lines) < max_lines:
            left_lines.append("")
        while len(right_lines) < max_lines:
            right_lines.append("")
        
        print(f"DEBUG: After padding - Left lines: {len(left_lines)}, Right lines: {len(right_lines)}")
        
        # Print both columns side by side
        for j in range(max_lines):
            left_line = left_lines[j] if j < len(left_lines) else ""
            right_line = right_lines[j] if j < len(right_lines) else ""
            print(f"{left_line:<40} {right_line}")
        
        print()  # Space between date pairs

# EXCEL EXPORT WITH UPDATED COLUMN LAYOUT
def export_to_excel(schedule):
    """Export schedule to Excel in simple, working format"""
    if not EXCEL_AVAILABLE:
        print("❌ openpyxl not installed. Install with: pip install openpyxl")
        return
        
    try:
        print("Creating Excel file with clean formats...")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create first half sheet
        create_simple_schedule_sheet(wb, schedule, "first")
        
        # Create second half sheet  
        create_simple_schedule_sheet(wb, schedule, "second")
        
        # Create grid sheet - full matrix format
        ws_grid = wb.create_sheet("Schedule Grid - All 22 Dates")
        ws_grid.sheet_view.showGridLines = False
        
        # Group games by date, time, and field
        games_by_date = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
        for game in schedule:
            games_by_date[game['date']][game['time']][game['field']] = game
        
        # Set up headers
        ws_grid['A1'] = "Division One - Fall 2025 Schedule"
        ws_grid['A1'].font = Font(size=16, bold=True)
        ws_grid.merge_cells('A1:S1')
        ws_grid['A1'].alignment = Alignment(horizontal='center')
        
        ws_grid['A2'] = "All games played at the Saddlebrook Complex"
        ws_grid['A2'].font = Font(size=12)
        ws_grid.merge_cells('A2:S2')
        ws_grid['A2'].alignment = Alignment(horizontal='center')
        
        # Column headers
        ws_grid['A4'] = "Date"
        ws_grid['B4'] = ""
        
        # 9:00 AM headers
        col = 3
        ws_grid.merge_cells(f'{get_column_letter(col)}3:{get_column_letter(col+8)}3')
        ws_grid[f'{get_column_letter(col)}3'] = "9:00"
        ws_grid[f'{get_column_letter(col)}3'].font = Font(bold=True)
        ws_grid[f'{get_column_letter(col)}3'].alignment = Alignment(horizontal='center')
        
        for field_num in [1, 2, 4]:  # Using 1,2,4 to match your image
            ws_grid[f'{get_column_letter(col)}4'] = "Field"
            ws_grid[f'{get_column_letter(col+1)}4'] = "Away"
            ws_grid[f'{get_column_letter(col+2)}4'] = "Home"
            col += 3
        
        # 10:30 AM headers  
        ws_grid.merge_cells(f'{get_column_letter(col)}3:{get_column_letter(col+8)}3')
        ws_grid[f'{get_column_letter(col)}3'] = "10:30"
        ws_grid[f'{get_column_letter(col)}3'].font = Font(bold=True)
        ws_grid[f'{get_column_letter(col)}3'].alignment = Alignment(horizontal='center')
        
        for field_num in [1, 2, 4]:
            ws_grid[f'{get_column_letter(col)}4'] = "Field"
            ws_grid[f'{get_column_letter(col+1)}4'] = "Away"
            ws_grid[f'{get_column_letter(col+2)}4'] = "Home"
            col += 3
        
        # Data rows
        current_row = 5
        sorted_dates = sorted(games_by_date.keys())
        
        for date in sorted_dates:
            # Date column
            day_name = date.strftime('%a')
            date_str = date.strftime('%m-%d')
            ws_grid[f'A{current_row}'] = date_str
            ws_grid[f'B{current_row}'] = day_name
            
            col = 3
            
            # 9:00 AM games
            for field_name in ['Field 1', 'Field 2', 'Field 3']:  # Changed to 1,2,3
                game = games_by_date[date]['9:00 AM'].get(field_name)
                if game:
                    field_num = field_name.split()[-1]
                    ws_grid[f'{get_column_letter(col)}{current_row}'] = field_num
                    ws_grid[f'{get_column_letter(col+1)}{current_row}'] = game['away']
                    ws_grid[f'{get_column_letter(col+2)}{current_row}'] = game['home']
                col += 3
            
            # 10:30 AM games  
            for field_name in ['Field 1', 'Field 2', 'Field 3']:  # Changed to 1,2,3
                game = games_by_date[date]['10:30 AM'].get(field_name)
                if game:
                    field_num = field_name.split()[-1]
                    ws_grid[f'{get_column_letter(col)}{current_row}'] = field_num
                    ws_grid[f'{get_column_letter(col+1)}{current_row}'] = game['away'] 
                    ws_grid[f'{get_column_letter(col+2)}{current_row}'] = game['home']
                col += 3
            
            current_row += 1
        
        # Add blackout row
        blackout_date = datetime.date(2025, 10, 17)
        if blackout_date not in games_by_date:
            ws_grid[f'A{current_row}'] = blackout_date.strftime('%m-%d')
            ws_grid[f'B{current_row}'] = blackout_date.strftime('%a')
            ws_grid.merge_cells(f'C{current_row}:S{current_row}')
            ws_grid[f'C{current_row}'] = "NO GAMES - FIELDS CLOSED FOR VILLAGES TOURNAMENT"
            ws_grid[f'C{current_row}'].font = Font(italic=True)
            ws_grid[f'C{current_row}'].alignment = Alignment(horizontal='center')
        
        # Auto-size columns
        for col in range(1, 20):
            ws_grid.column_dimensions[get_column_letter(col)].width = 12
        
        filename = "villages_d1_softball_fall_2025_schedule_CLEAN.xlsx"
        wb.save(filename)
        print(f"✅ Excel file created: {filename}")
        
    except Exception as e:
        print(f"❌ Excel error: {e}")

def create_simple_schedule_sheet(wb, schedule, half):
    """Create a simple, working schedule sheet"""
    
    # Split schedule
    game_dates_only = sorted(set(game['date'] for game in schedule))
    split_point = len(game_dates_only) // 2
    
    if half == "first":
        dates = game_dates_only[:split_point] 
        sheet_name = "Schedule - First Half"
    else:
        dates = game_dates_only[split_point:]
        sheet_name = "Schedule - Second Half"
    
    games = [game for game in schedule if game['date'] in dates]
    
    # Create worksheet
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    
    # Set column widths - updated for new layout: A(new), B(margin), C(wed), D(spacer), E(fri), F(margin)
    ws.column_dimensions['A'].width = 4   # New inserted column
    ws.column_dimensions['B'].width = 4   # Left margin
    ws.column_dimensions['C'].width = 42  # Wednesday content 
    ws.column_dimensions['D'].width = 4   # Spacer
    ws.column_dimensions['E'].width = 42  # Friday content
    ws.column_dimensions['F'].width = 4   # Right margin
    
    # Title - updated for new layout
    ws['C3'] = f"Villages D1 Softball - Fall 2025"
    ws['C3'].font = Font(size=18, bold=True, color="1F4E79")
    ws.merge_cells('C3:E3')
    ws['C3'].alignment = Alignment(horizontal='center')
    ws['C3'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    ws['C4'] = f"Schedule - {half.title()} Half"
    ws['C4'].font = Font(size=14, italic=True, color="4472C4")  
    ws.merge_cells('C4:E4')
    ws['C4'].alignment = Alignment(horizontal='center')
    
    # Get all dates with special events
    all_dates = get_all_dates_simple(games, dates, half)
    
    # Group games by date
    games_by_date = defaultdict(list)
    for game in games:
        games_by_date[game['date']].append(game)
    
    # Create weekly pairs
    wednesdays = [(date, info) for date, info in all_dates if date.weekday() == 2]
    fridays = [(date, info) for date, info in all_dates if date.weekday() == 4]
    
    weekly_pairs = []
    for wed_date, wed_info in wednesdays:
        target_friday = wed_date + datetime.timedelta(days=2)
        fri_match = None
        for fri_date, fri_info in fridays:
            if fri_date == target_friday:
                fri_match = (fri_date, fri_info)
                break
        weekly_pairs.append(((wed_date, wed_info), fri_match))
    
    current_row = 6
    
    # Process weekly pairs - updated for new layout
    for wed_info, fri_info in weekly_pairs:
        start_row = current_row
        
        # Wednesday column - now column C
        if wed_info:
            current_row = add_simple_date_section(ws, wed_info[0], wed_info[1], games_by_date, current_row, 'C')
        
        # Friday column - now column E
        temp_row = start_row
        if fri_info:
            temp_row = add_simple_date_section(ws, fri_info[0], fri_info[1], games_by_date, temp_row, 'E')
        
        current_row = max(current_row, temp_row) + 1
    
    # Add simple borders - CONNECTED version
    try:
        end_row = current_row + 1
        
        # Apply borders that actually connect properly
        outer_border = Side(style='medium', color='2F5597')
        inner_border = Side(style='medium', color='4472C4')
        inner_end = end_row - 1
        
        # Outer border - create complete connected rectangle
        for r in range(1, end_row + 1):
            for c in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws[f'{c}{r}']
                borders = {}
                
                # Add border sides that touch the perimeter
                if r == 1:  # Top edge
                    borders['top'] = outer_border
                if r == end_row:  # Bottom edge
                    borders['bottom'] = outer_border
                if c == 'A':  # Left edge
                    borders['left'] = outer_border
                if c == 'F':  # Right edge
                    borders['right'] = outer_border
                    
                if borders:
                    cell.border = Border(**borders)
        
        # Inner border - create complete connected rectangle
        for r in range(3, inner_end + 1):
            for c in ['C', 'D', 'E']:
                cell = ws[f'{c}{r}']
                borders = {}
                
                # Add border sides that touch the inner perimeter
                if r == 3:  # Top edge of inner
                    borders['top'] = inner_border
                if r == inner_end:  # Bottom edge of inner
                    borders['bottom'] = inner_border
                if c == 'C':  # Left edge of inner
                    borders['left'] = inner_border
                if c == 'E':  # Right edge of inner
                    borders['right'] = inner_border
                    
                if borders:
                    # Preserve any existing outer borders and add inner borders
                    existing = cell.border
                    cell.border = Border(
                        top=borders.get('top') or existing.top,
                        bottom=borders.get('bottom') or existing.bottom,
                        left=borders.get('left') or existing.left,
                        right=borders.get('right') or existing.right
                    )
                    
    except:
        pass  # Skip borders if they fail


def add_simple_date_section(ws, date, date_info, games_by_date, start_row, col):
    """Add a simple date section"""
    current_row = start_row
    
    # Date header
    ws[f'{col}{current_row}'] = date.strftime('%A, %B %d')
    ws[f'{col}{current_row}'].font = Font(size=12, bold=True, color="2F5597")
    ws[f'{col}{current_row}'].fill = PatternFill(start_color="E6F1FF", end_color="E6F1FF", fill_type="solid")
    current_row += 1
    
    # Special events
    if date_info.get('special_event'):
        ws[f'{col}{current_row}'] = date_info['special_event']
        ws[f'{col}{current_row}'].font = Font(size=10, italic=True, color="C55A11")
        ws[f'{col}{current_row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        current_row += 1
    elif date in games_by_date:
        # Sort games by time
        date_games = sorted(games_by_date[date], 
                          key=lambda g: (0 if '9:00' in g['time'] else 1, int(g['field'].split()[-1])))
        
        for game in date_games:
            field_num = game['field'].split()[-1]
            time_str = game['time'].replace(' AM', '').replace(' PM', '')
            game_text = f"{time_str} {game['away']} at {game['home']} on {field_num}"
            
            ws[f'{col}{current_row}'] = game_text
            ws[f'{col}{current_row}'].font = Font(size=10, color="333333")
            current_row += 1
    else:
        ws[f'{col}{current_row}'] = "No games scheduled"
        ws[f'{col}{current_row}'].font = Font(size=10, italic=True, color="888888")
        current_row += 1
    
    return current_row

def get_all_dates_simple(games, date_range, half):
    """Get all dates with special events - simple version"""
    special_events = {
        datetime.date(2025, 9, 3): {'special_event': 'PRACTICE DAY - Team Preparation'},
        datetime.date(2025, 10, 17): {'special_event': 'BLACKOUT - Villages Tournament'},
        datetime.date(2025, 11, 26): {'special_event': 'RAIN-OUT MAKEUP OR POST SEASON PLAYOFFS'},
        datetime.date(2025, 11, 28): {'special_event': 'TURKEY BOWL - THANKSGIVING TOURNAMENT'},
        datetime.date(2025, 12, 3): {'special_event': 'POST SEASON PLAYOFFS'},
        datetime.date(2025, 12, 5): {'special_event': 'POST SEASON PLAYOFFS'},
        datetime.date(2025, 12, 10): {'special_event': 'POST SEASON PLAYOFFS - IF NECESSARY'},
        datetime.date(2025, 12, 12): {'special_event': 'POST SEASON PLAYOFFS - IF NECESSARY'},
    }
    
    all_dates = []
    
    # Add game dates
    for date in date_range:
        all_dates.append((date, {}))
    
    # Add special events
    min_date = min(date_range)
    max_date = max(date_range)
    
    for special_date, info in special_events.items():
        include = False
        
        if half == "first":
            if special_date == datetime.date(2025, 9, 3) and any(d.month == 9 for d in date_range):
                include = True
            elif min_date <= special_date <= max_date:
                include = True
        else:  # second half
            if min_date <= special_date <= max_date:
                include = True
            elif special_date.month >= 11:
                include = True
        
        if include:
            all_dates.append((special_date, info))
    
    # Remove duplicates and sort
    seen = set()
    unique_dates = []
    for date, info in all_dates:
        if date not in seen:
            seen.add(date)
            unique_dates.append((date, info))
    
    unique_dates.sort(key=lambda x: x[0])
    return unique_dates

def export_csv(schedule):
    """Export to CSV"""
    filename = "villages_d1_softball_fall_2025_data.csv"
    
    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Date', 'Round', 'Home Team', 'Away Team', 'Field', 'Time'])
        
        for game in sorted(schedule, key=lambda x: (x['round'], x['field'], x['time'])):
            writer.writerow([
                game['date'].strftime('%m/%d/%Y'),
                game['round'],
                game['home'],
                game['away'],
                game['field'],
                game['time']
            ])
    
    print(f"✅ Data exported to {filename}")

def main():
    """Generate optimized schedule"""
    print("Villages D1 Softball Schedule Generator")
    print("Deterministic Valid Base + Intelligent Consecutive Optimization")
    print("=" * 70)
    
    # Generate guaranteed valid base
    base_schedule = generate_guaranteed_valid_schedule()
    
    # Verify it's valid
    violations = verify_schedule_validity(base_schedule)
    if violations:
        print("❌ Base schedule has violations - this should never happen!")
        for violation in violations:
            print(f"  - {violation}")
        return
    
    print("✅ Base schedule verified - all hard constraints met")
    
    # Optimize consecutive games
    optimized_schedule, violation_details = optimize_consecutive_games_intelligently(base_schedule, max_iterations=15000)
    
    # Final verification
    final_violations = verify_schedule_validity(optimized_schedule)
    if final_violations:
        print("❌ CRITICAL ERROR: Optimization broke hard constraints!")
        for violation in final_violations:
            print(f"  - {violation}")
        return
    
    print("✅ Final schedule verified - all hard constraints still met")
    
    # *** MAIN SCHEDULE OUTPUT ***
    print_professional_schedule(optimized_schedule)
    
    # Export to Excel and CSV
    export_to_excel(optimized_schedule)
    export_csv(optimized_schedule)
    
    # Technical analysis (can be skipped for final output)
    print(f"\n\n{'='*60}")
    print("TECHNICAL ANALYSIS (for verification)")
    print(f"{'='*60}")
    
    teams_with_violations = analyze_final_schedule(optimized_schedule, violation_details)
    
    print(f"\n🎉 OPTIMIZATION SUMMARY:")
    print(f"Teams with >4 consecutive games: {teams_with_violations}")
    print(f"Hard constraints: GUARANTEED VALID")

if __name__ == "__main__":
    main()