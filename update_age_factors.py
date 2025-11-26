#!/usr/bin/env python3
"""
Update age factors with better scale:
- Under 60 = +5 (significantly younger)
- 60-64 = +3 (young side)
- 65-69 = +1 (around average - 63)
- 70-74 = 0 (neutral)
- 75-79 = -2 (older)
- 80+ = -4 (significantly older)
"""

import pandas as pd
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# Recalculate age factors with new scale
print("\nRecalculating age factors...")
age_factors = []
for idx, row in df.iterrows():
    age = row['Age']
    if pd.notna(age):
        if age < 60:
            age_factors.append(5)
        elif age <= 64:
            age_factors.append(3)
        elif age <= 69:
            age_factors.append(1)
        elif age <= 74:
            age_factors.append(0)
        elif age <= 79:
            age_factors.append(-2)
        else:  # 80+
            age_factors.append(-4)
    else:
        age_factors.append(0)

df['Age_Factor'] = age_factors

# Recalculate Ranking_Score
print("Recalculating Ranking_Score...")
ranking_scores = []
for idx, row in df.iterrows():
    base = row['Weighted_convBA+'] if pd.notna(row['Weighted_convBA+']) else 0
    defense = row['Def_Spectrum_Pts']
    age = row['Age_Factor']
    manual = row['Manual_Adj']
    
    if base > 0:  # Only rank players with stats
        score = round(base + defense + age + manual, 1)
        ranking_scores.append(score)
    else:
        ranking_scores.append(None)

df['Ranking_Score'] = ranking_scores

# Sort by Ranking_Score (descending)
df = df.sort_values('Ranking_Score', ascending=False, na_position='last')

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply convBA formatting
print("Re-applying .000 format to convBA columns...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns (H, I, J)
for col_idx in [8, 9, 10]:  # convBA_F25, S25, W25
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Updated Age_Factor scale:")
print(f"    Under 60 = +5")
print(f"    60-64 = +3")
print(f"    65-69 = +1 (around average)")
print(f"    70-74 = 0")
print(f"    75-79 = -2")
print(f"    80+ = -4")
print(f"  Recalculated Ranking_Score")
print(f"  Re-sorted by Ranking_Score")
print(f"  Backup: {backup_file}")

print(f"\nTop 10 players:")
print(df[['FirstName', 'LastName', 'Position', 'Age', 'Weighted_convBA+', 'Def_Spectrum_Pts', 'Age_Factor', 'Ranking_Score']].head(10).to_string(index=False))

