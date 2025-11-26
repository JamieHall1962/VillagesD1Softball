#!/usr/bin/env python3
"""
Update to Career Availability %
- Calculate across all seasons (or last 20 seasons)
- Shows historical reliability pattern
"""

import pandas as pd
import sqlite3
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'
DATABASE_PATH = 'softball_stats.db'
MAX_SEASONS = 20  # Look back maximum 20 seasons

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# Connect to database
print(f"\nCalculating career availability % (last {MAX_SEASONS} seasons)...")
conn = sqlite3.connect(DATABASE_PATH)

def get_career_availability(pid):
    """Calculate career availability percentage"""
    cursor = conn.cursor()
    
    # Get all teams the player has been on (most recent first)
    teams_query = """
        SELECT DISTINCT b.TeamNumber, t.LongTeamName
        FROM batting_stats b
        INNER JOIN Teams t ON b.TeamNumber = t.TeamNumber
        WHERE b.PlayerNumber = ?
        ORDER BY t.LongTeamName DESC
        LIMIT ?
    """
    cursor.execute(teams_query, (pid, MAX_SEASONS))
    teams = cursor.fetchall()
    
    if not teams:
        return None, 0, 0
    
    total_games_played = 0
    total_games_possible = 0
    
    for team in teams:
        team_number = team[0]
        
        # Count games the team played
        cursor.execute("""
            SELECT COUNT(DISTINCT GameNumber)
            FROM game_stats
            WHERE TeamNumber = ?
        """, (team_number,))
        team_games = cursor.fetchone()[0]
        
        # Count games player appeared in for this team
        cursor.execute("""
            SELECT COUNT(DISTINCT GameNumber)
            FROM batting_stats
            WHERE PlayerNumber = ?
            AND TeamNumber = ?
        """, (pid, team_number))
        player_games = cursor.fetchone()[0]
        
        total_games_played += player_games
        total_games_possible += team_games
    
    if total_games_possible > 0:
        availability_pct = round((total_games_played / total_games_possible) * 100, 1)
        return availability_pct, total_games_played, total_games_possible
    else:
        return None, 0, 0

# Calculate for each player
career_availability_pcts = []
career_games_played = []
career_games_possible = []

for idx, row in df.iterrows():
    pid = row['PID']
    
    if pd.notna(pid):
        pid = int(pid)
        avail_pct, games_played, games_possible = get_career_availability(pid)
        career_availability_pcts.append(avail_pct)
        career_games_played.append(games_played)
        career_games_possible.append(games_possible)
    else:
        career_availability_pcts.append(None)
        career_games_played.append(0)
        career_games_possible.append(0)

conn.close()

# Replace the F25-only columns with career columns
df['Career_Games_Played'] = career_games_played
df['Career_Games_Possible'] = career_games_possible
df['Career_Availability_Pct'] = career_availability_pcts

# Drop the old F25-specific columns
if 'Games_Played_F25' in df.columns:
    df = df.drop(columns=['Games_Played_F25', 'Team_Games_F25', 'Availability_Pct'])

# Reorder columns
new_order = [
    'PID',
    'FirstName',
    'LastName',
    'Position',
    'Age',
    'PA',
    'HR',
    'Career_Games_Played',
    'Career_Games_Possible',
    'Career_Availability_Pct',
    'Win_Pct',
    'Career_Wins',
    'Career_Losses',
    'convBA_F25',
    'convBA_S25',
    'convBA_W25',
    'convBA+_F25',
    'convBA+_S25',
    'convBA+_W25',
    'Weighted_convBA+',
    'Def_Spectrum_Pts',
    'Age_Factor',
    'Manual_Adj',
    'Ranking_Score',
    'Availability'
]

df = df[new_order]

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns (N, O, P)
for col_idx in [14, 15, 16]:  # convBA_F25, S25, W25
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column (K)
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)  # Win_Pct column
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Replaced F25-only with Career Availability (last {MAX_SEASONS} seasons)")
print(f"  Columns: Career_Games_Played, Career_Games_Possible, Career_Availability_Pct")
print(f"  Backup: {backup_file}")

print(f"\nTop 10 with Career Availability %:")
display_cols = ['FirstName', 'LastName', 'Career_Games_Played', 'Career_Games_Possible', 'Career_Availability_Pct', 'Ranking_Score']
print(df[display_cols].head(10).to_string(index=False))

print(f"\nHighest Career Availability (90%+):")
high_avail = df[df['Career_Availability_Pct'] >= 90.0].sort_values('Career_Availability_Pct', ascending=False)
if len(high_avail) > 0:
    print(f"  {len(high_avail)} players with 90%+ career availability")
    print(high_avail[['FirstName', 'LastName', 'Career_Games_Played', 'Career_Games_Possible', 'Career_Availability_Pct']].head(15).to_string(index=False))

print(f"\nLowest Career Availability (<75%):")
low_avail = df[df['Career_Availability_Pct'] < 75.0].sort_values('Career_Availability_Pct')
if len(low_avail) > 0:
    print(f"  {len(low_avail)} players with <75% career availability")
    print(low_avail[['FirstName', 'LastName', 'Career_Games_Played', 'Career_Games_Possible', 'Career_Availability_Pct']].head(10).to_string(index=False))

