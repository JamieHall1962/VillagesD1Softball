#!/usr/bin/env python3
"""
Update Defensive Spectrum points to more balanced scale:
SS: 5, LC/RC: 4, MI: 3.5, LF: 2.5, 3B: 2, RF/2B: 1.5, P: 1, 1B: 0.5, C: 0
"""

import pandas as pd
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# New defensive spectrum (more balanced)
NEW_DEFENSIVE_SPECTRUM = {
    'SS': 5,
    'LC': 4,
    'RC': 4,
    'MI': 3.5,
    'LF': 2.5,
    '3B': 2,
    'RF': 1.5,
    '2B': 1.5,
    'P': 1,
    '1B': 0.5,
    'C': 0
}

print("\nRecalculating Def_Spectrum_Pts with new scale...")

# Recalculate defensive spectrum points
def_points = []
for idx, row in df.iterrows():
    pos = row['Position']
    if pd.notna(pos) and pos in NEW_DEFENSIVE_SPECTRUM:
        def_points.append(NEW_DEFENSIVE_SPECTRUM[pos])
    else:
        def_points.append(0)  # Unknown position

df['Def_Spectrum_Pts'] = def_points

# Recalculate Ranking_Score
print("Recalculating Ranking_Score...")
ranking_scores = []
for idx, row in df.iterrows():
    base = row['Weighted_convBA+'] if pd.notna(row['Weighted_convBA+']) else 0
    defense = row['Def_Spectrum_Pts']
    age = row['Age_Factor']
    manual = row['Manual_Adj']
    avail_adj = row['Availability_Adj']
    
    if base > 0:  # Only rank players with stats
        score = round(base + defense + age + manual + avail_adj, 1)
        ranking_scores.append(score)
    else:
        ranking_scores.append(None)

df['Ranking_Score'] = ranking_scores

# Sort by Ranking_Score (descending)
df = df.sort_values('Ranking_Score', ascending=False, na_position='last')

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns (N, O, P)
for col_idx in [14, 15, 16]:  # convBA_F25, S25, W25
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column (K)
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)  # Win_Pct column
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Updated Defensive Spectrum scale:")
print(f"    SS: 5, LC/RC: 4, MI: 3.5, LF: 2.5")
print(f"    3B: 2, RF/2B: 1.5, P: 1, 1B: 0.5, C: 0")
print(f"  Recalculated Ranking_Score")
print(f"  Re-sorted by Ranking_Score")
print(f"  Backup: {backup_file}")

print(f"\nTop 15 with updated defensive spectrum:")
display_cols = ['FirstName', 'LastName', 'Position', 'Def_Spectrum_Pts', 'Weighted_convBA+', 'Ranking_Score']
print(df[display_cols].head(15).to_string(index=False))

print(f"\nPositional breakdown of top 30:")
top30 = df.head(30)
pos_counts = top30['Position'].value_counts()
print(pos_counts.to_string())

