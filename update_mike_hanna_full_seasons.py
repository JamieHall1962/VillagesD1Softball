#!/usr/bin/env python3
"""
Update Mike Hanna to use last 3 FULL seasons (ignore injury/partial seasons)
Filter out seasons with < 20 PA
"""

import pandas as pd
import sqlite3
from datetime import datetime
import shutil
from openpyxl import load_workbook
import numpy as np

EXCEL_FILE = 'w26rankings.xlsx'
DATABASE_PATH = 'softball_stats.db'
MIN_PA_FULL_SEASON = 20  # Minimum PA to count as a "full" season

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')

conn = sqlite3.connect(DATABASE_PATH)
pid = 530  # Mike Hanna

# Get all seasons with stats
query = """
    SELECT t.LongTeamName,
           SUM(PA) as PA,
           SUM(H) as H,
           SUM(BB) as BB,
           SUM(SF) as SF,
           SUM(OE) as OE,
           SUM([2B]) as Doubles,
           SUM([3B]) as Triples,
           SUM(HR) as HR
    FROM batting_stats bs
    INNER JOIN Teams t ON bs.TeamNumber = t.TeamNumber
    WHERE bs.PlayerNumber = ?
    GROUP BY bs.TeamNumber
    ORDER BY t.LongTeamName DESC
"""

cursor = conn.cursor()
cursor.execute(query, (pid,))
seasons_data = cursor.fetchall()

print(f"\nAll seasons for Mike Hanna:")
for s in seasons_data:
    season_code = s[0].split()[-1] if s[0] else ''
    pa = s[1]
    print(f"  {season_code}: {pa} PA")

# Calculate convBA+ for each FULL season (PA >= 20)
def get_season_sort_value(season_str):
    if not season_str or len(season_str) < 3:
        return 0
    season_type = season_str[0]
    try:
        year = int(season_str[1:])
    except:
        return 0
    season_order = {'F': 3, 'S': 2, 'W': 1}.get(season_type, 0)
    return year * 10 + season_order

season_convba_plus = []
for season_row in seasons_data:
    team_name = season_row[0]
    pa = season_row[1]
    
    # FILTER: Only include full seasons (PA >= 20)
    if pa < MIN_PA_FULL_SEASON:
        season_code = team_name.split()[-1] if team_name else ''
        print(f"  Skipping {season_code} - only {pa} PA (injury/partial season)")
        continue
    
    h = season_row[2]
    bb = season_row[3]
    sf = season_row[4]
    oe = season_row[5]
    doubles = season_row[6]
    triples = season_row[7]
    hr = season_row[8]
    
    total_bases = h + doubles + (2 * triples) + (3 * hr)
    conv_ba = (((4*(h+bb)+total_bases)/pa)/0.305*0.25)/10
    
    season_code = team_name.split()[-1] if team_name else None
    if season_code:
        # Get league average
        league_query = """
            SELECT SUM(PA) as PA, SUM(H) as H, SUM(BB) as BB,
                   SUM([2B]) as Doubles, SUM([3B]) as Triples, SUM(HR) as HR
            FROM batting_stats bs
            INNER JOIN Teams t ON bs.TeamNumber = t.TeamNumber
            WHERE t.LongTeamName LIKE ?
        """
        cursor.execute(league_query, (f'% {season_code}',))
        league_row = cursor.fetchone()
        
        if league_row and league_row[0]:
            l_pa, l_h, l_bb, l_doubles, l_triples, l_hr = league_row
            l_tb = l_h + l_doubles + (2 * l_triples) + (3 * l_hr)
            league_conv_ba = (((4*(l_h+l_bb)+l_tb)/l_pa)/0.305*0.25)/10
            
            conv_ba_plus = (conv_ba / league_conv_ba) * 100
            season_convba_plus.append({
                'season': season_code,
                'convBA+': conv_ba_plus,
                'convBA': conv_ba,
                'sort_value': get_season_sort_value(season_code),
                'PA': pa,
                'HR': hr
            })

# Sort and take last 3 FULL seasons
season_convba_plus.sort(key=lambda x: x['sort_value'], reverse=True)
last_3_seasons = season_convba_plus[:3]

print(f"\nLast 3 FULL seasons used:")
for s in last_3_seasons:
    print(f"  {s['season']}: {s['PA']} PA, {s['convBA+']:.1f} convBA+")

# Calculate weighted average
base_weights = [0.50, 0.30, 0.20]
total_weight = sum(base_weights[:len(last_3_seasons)])
weighted_sum = sum(s['convBA+'] * (base_weights[i] / total_weight) 
                  for i, s in enumerate(last_3_seasons))
weighted_convba_plus = round(weighted_sum, 1)
seasons_used = ', '.join([s['season'] for s in last_3_seasons])

print(f"\nWeighted convBA+: {weighted_convba_plus}")

# Update Mike Hanna's row
hanna_idx = df[df['PID'] == 530].index[0]
df.at[hanna_idx, 'Weighted_convBA+'] = weighted_convba_plus
df.at[hanna_idx, 'Seasons_Used'] = seasons_used

# Update PA and HR to most recent full season
if last_3_seasons:
    df.at[hanna_idx, 'PA'] = last_3_seasons[0]['PA']
    df.at[hanna_idx, 'HR'] = last_3_seasons[0]['HR']

# Recalculate confidence for Hanna
most_recent = last_3_seasons[0]['season'] if last_3_seasons else ''
if most_recent in ['F25', 'W25']:
    recency = 1.0
elif most_recent == 'S25':
    recency = 0.8
elif most_recent in ['F24', 'W24']:
    recency = 0.6
else:
    recency = 0.4

total_pa = sum(s['PA'] for s in last_3_seasons)
if total_pa >= 150:
    sample_size = 1.0
elif total_pa >= 100:
    sample_size = 0.75
elif total_pa >= 50:
    sample_size = 0.5
else:
    sample_size = 0.25

num_seasons = len(last_3_seasons)
if num_seasons >= 3:
    seasons_score = 1.0
elif num_seasons == 2:
    seasons_score = 0.67
else:
    seasons_score = 0.33

if num_seasons >= 2:
    variance = np.std([s['convBA+'] for s in last_3_seasons])
    if variance < 15:
        consistency = 1.0
    elif variance < 30:
        consistency = 0.75
    else:
        consistency = 0.5
else:
    consistency = 0.75

confidence_pct = round((recency * 0.40 + sample_size * 0.30 + seasons_score * 0.20 + consistency * 0.10) * 100, 1)
if confidence_pct >= 90:
    confidence = 'A'
elif confidence_pct >= 80:
    confidence = 'B'
elif confidence_pct >= 70:
    confidence = 'C'
elif confidence_pct >= 60:
    confidence = 'D'
else:
    confidence = 'F'

df.at[hanna_idx, 'Confidence'] = confidence
df.at[hanna_idx, 'Confidence_Pct'] = confidence_pct

# Recalculate ranking score
base = weighted_convba_plus
defense = df.at[hanna_idx, 'Def_Spectrum_Pts']
age_factor = df.at[hanna_idx, 'Age_Factor']
manual = df.at[hanna_idx, 'Manual_Adj']
avail_adj = df.at[hanna_idx, 'Availability_Adj']

ranking_score = round(base + defense + age_factor + manual + avail_adj, 1)
df.at[hanna_idx, 'Ranking_Score'] = ranking_score

# Sort by Ranking_Score
df = df.sort_values('Ranking_Score', ascending=False, na_position='last')

conn.close()

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']
for col_idx in [14, 15, 16]:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'
wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Updated Mike Hanna with last 3 FULL seasons")
print(f"  New weighted convBA+: {weighted_convba_plus}")
print(f"  New ranking score: {ranking_score}")
print(f"  New confidence: {confidence} ({confidence_pct}%)")
print(f"  Backup: {backup_file}")

# Show his new rank
hanna_rank = df[df['PID'] == 530].index[0] + 1
print(f"\nMike Hanna new rank: #{hanna_rank}")

