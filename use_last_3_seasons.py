#!/usr/bin/env python3
"""
Recalculate weighted convBA+ using each player's LAST 3 SEASONS PLAYED
- Find all seasons player has data for
- Sort by most recent
- Take last 3 seasons played
- Weight: Most recent 50%, 2nd 30%, 3rd 20%
- Normalize if fewer than 3 seasons
"""

import pandas as pd
import sqlite3
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = 'w26rankings.xlsx'
DATABASE_PATH = 'softball_stats.db'

# Create backup
backup_file = f"w26rankings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
print(f"Creating backup: {backup_file}")
shutil.copy(EXCEL_FILE, backup_file)

# Load rankings
print(f"Loading {EXCEL_FILE}...")
df = pd.read_excel(EXCEL_FILE, sheet_name='rankings')
print(f"  Found {len(df)} players")

# Connect to database
print("\nGetting last 3 seasons played for each player...")
conn = sqlite3.connect(DATABASE_PATH)

def get_season_sort_value(season_str):
    """Convert season string (like 'F25', 'W24') to sortable value"""
    # Format: [Season Type][Year]
    # F=Fall, S=Spring/Summer, W=Winter
    # Higher = more recent
    if not season_str or len(season_str) < 3:
        return 0
    
    season_type = season_str[0]
    try:
        year = int(season_str[1:])
    except:
        return 0
    
    # Convert to sortable: year * 10 + season_order
    # F (Fall) = 3, S (Spring/Summer) = 2, W (Winter) = 1
    season_order = {'F': 3, 'S': 2, 'W': 1}.get(season_type, 0)
    
    return year * 10 + season_order

def get_last_3_seasons_convba_plus(pid):
    """Get convBA+ for player's last 3 seasons played"""
    query = """
        SELECT t.LongTeamName,
               SUM(PA) as PA,
               SUM(H) as H,
               SUM(BB) as BB,
               SUM(SF) as SF,
               SUM(OE) as OE,
               SUM([2B]) as Doubles,
               SUM([3B]) as Triples,
               SUM(HR) as HR
        FROM batting_stats bs
        INNER JOIN Teams t ON bs.TeamNumber = t.TeamNumber
        WHERE bs.PlayerNumber = ?
        GROUP BY bs.TeamNumber
        ORDER BY t.LongTeamName DESC
    """
    
    cursor = conn.cursor()
    cursor.execute(query, (pid,))
    seasons_data = cursor.fetchall()
    
    if not seasons_data:
        return []
    
    # Calculate convBA+ for each season
    season_convba_plus = []
    
    for season_row in seasons_data:
        team_name = season_row[0]
        pa = season_row[1]
        h = season_row[2]
        bb = season_row[3]
        sf = season_row[4]
        oe = season_row[5]
        doubles = season_row[6]
        triples = season_row[7]
        hr = season_row[8]
        
        if pa > 0:
            # Calculate convBA
            total_bases = h + doubles + (2 * triples) + (3 * hr)
            conv_ba = (((4*(h+bb)+total_bases)/pa)/0.305*0.25)/10
            
            # Get league average for this season
            season_code = team_name.split()[-1] if team_name else None
            
            if season_code:
                # Get league average for this season
                league_query = """
                    SELECT 
                        SUM(PA) as PA,
                        SUM(H) as H,
                        SUM(BB) as BB,
                        SUM([2B]) as Doubles,
                        SUM([3B]) as Triples,
                        SUM(HR) as HR
                    FROM batting_stats bs
                    INNER JOIN Teams t ON bs.TeamNumber = t.TeamNumber
                    WHERE t.LongTeamName LIKE ?
                """
                cursor.execute(league_query, (f'% {season_code}',))
                league_row = cursor.fetchone()
                
                if league_row and league_row[0]:
                    l_pa, l_h, l_bb, l_doubles, l_triples, l_hr = league_row
                    l_tb = l_h + l_doubles + (2 * l_triples) + (3 * l_hr)
                    league_conv_ba = (((4*(l_h+l_bb)+l_tb)/l_pa)/0.305*0.25)/10
                    
                    # Calculate convBA+
                    conv_ba_plus = (conv_ba / league_conv_ba) * 100
                    
                    season_convba_plus.append({
                        'season': season_code,
                        'convBA+': conv_ba_plus,
                        'sort_value': get_season_sort_value(season_code)
                    })
    
    # Sort by most recent and take last 3
    season_convba_plus.sort(key=lambda x: x['sort_value'], reverse=True)
    return season_convba_plus[:3]

# Calculate weighted convBA+ for each player
weighted_convba_plus_list = []
seasons_used_list = []

for idx, row in df.iterrows():
    pid = row['PID']
    
    if pd.notna(pid):
        pid = int(pid)
        last_3_seasons = get_last_3_seasons_convba_plus(pid)
        
        if last_3_seasons:
            # Apply weights: 50%, 30%, 20%
            base_weights = [0.50, 0.30, 0.20]
            
            # Use only available seasons and normalize weights
            total_weight = sum(base_weights[:len(last_3_seasons)])
            
            weighted_sum = 0
            for i, season_data in enumerate(last_3_seasons):
                normalized_weight = base_weights[i] / total_weight
                weighted_sum += season_data['convBA+'] * normalized_weight
            
            weighted_convba_plus_list.append(round(weighted_sum, 1))
            
            # Track which seasons were used
            seasons_str = ', '.join([s['season'] for s in last_3_seasons])
            seasons_used_list.append(seasons_str)
        else:
            weighted_convba_plus_list.append(None)
            seasons_used_list.append('')
    else:
        weighted_convba_plus_list.append(None)
        seasons_used_list.append('')

conn.close()

# Update dataframe
df['Weighted_convBA+'] = weighted_convba_plus_list
df['Seasons_Used'] = seasons_used_list

# Recalculate Ranking_Score
print("Recalculating Ranking_Score...")
ranking_scores = []
for idx, row in df.iterrows():
    base = row['Weighted_convBA+'] if pd.notna(row['Weighted_convBA+']) else 0
    defense = row['Def_Spectrum_Pts']
    age = row['Age_Factor']
    manual = row['Manual_Adj']
    avail_adj = row['Availability_Adj']
    
    if base > 0:
        score = round(base + defense + age + manual + avail_adj, 1)
        ranking_scores.append(score)
    else:
        ranking_scores.append(None)

df['Ranking_Score'] = ranking_scores

# Sort by Ranking_Score
df = df.sort_values('Ranking_Score', ascending=False, na_position='last')

# Reorder columns - add Seasons_Used after Weighted_convBA+
cols = df.columns.tolist()
weighted_idx = cols.index('Weighted_convBA+')
cols.insert(weighted_idx + 1, cols.pop(cols.index('Seasons_Used')))
df = df[cols]

# Save
print(f"\nSaving {EXCEL_FILE}...")
with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='rankings', index=False)

# Re-apply formatting
print("Re-applying number formats...")
wb = load_workbook(EXCEL_FILE)
ws = wb['rankings']

# Format convBA columns
for col_idx in [14, 15, 16]:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '.000'

# Format Win_Pct column
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=11)
    if cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = '.000'

wb.save(EXCEL_FILE)

print(f"\n[DONE]")
print(f"  Now using each player's LAST 3 SEASONS PLAYED")
print(f"  Weights: Most recent 50%, 2nd 30%, 3rd 20%")
print(f"  Added 'Seasons_Used' column showing which seasons were used")
print(f"  Backup: {backup_file}")

print(f"\nTop 20 rankings with seasons used:")
display_cols = ['FirstName', 'LastName', 'Position', 'Weighted_convBA+', 'Seasons_Used', 'Ranking_Score']
print(df[display_cols].head(20).to_string(index=False))

